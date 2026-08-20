"""Тесты страницы «Остатки для FBO».

Главное, что здесь проверяется: ожидание (плановый выпуск по производственным
заданиям) не попадает в «Можно взять». Именно из-за него колонка «Доступно» в
МойСклад показывает больше, чем реально лежит на складе.
"""
from io import BytesIO
from unittest.mock import MagicMock, patch

from django.contrib.auth.models import User
from django.core.cache import cache
from django.test import Client, SimpleTestCase, TestCase
from openpyxl import load_workbook

from api.access import page_keys_for_path
from api.models import UserPageAccess
from api.views.fbo_stock import _build_data, _minimum_balances

STORE_HREF = 'https://api.moysklad.ru/api/remap/1.2/entity/store/store-1'
FOLDER_HREF = 'https://api.moysklad.ru/api/remap/1.2/entity/productfolder/folder-1'

# Реальные цифры из МойСклад на 20.08.2026: пробиотик ждёт выпуска 70 шт по
# производственному заданию, биотин — без ожидания, у геля резерв больше остатка.
STOCK_ROWS = [
    {
        'meta': {'href': 'https://api.moysklad.ru/api/remap/1.2/entity/product/probiotic'},
        'article': '06-03GP0800', 'code': '2-207', 'name': 'Пробиотик GastroPro для лошадей, 800г',
        'folder': {'name': 'GastroPro'}, 'stock': 237.0, 'reserve': 18.0, 'inTransit': 70.0, 'quantity': 289.0,
    },
    {
        'meta': {'href': 'https://api.moysklad.ru/api/remap/1.2/entity/product/biotin'},
        'article': '08-02EP0800', 'code': '2-156', 'name': 'Биотин ExterPro для лошадей, 800 г',
        'folder': {'name': 'ExterPro'}, 'stock': 213.0, 'reserve': 12.0, 'inTransit': 0.0, 'quantity': 201.0,
    },
    {
        'meta': {'href': 'https://api.moysklad.ru/api/remap/1.2/entity/product/gel'},
        'article': '01-14AP0250', 'code': '2-090', 'name': 'Юнифлекс гель ArtroPro, 250 мл',
        'folder': {'name': 'ArtroPro'}, 'stock': 10.0, 'reserve': 18.0, 'inTransit': 60.0, 'quantity': 52.0,
    },
    {
        'meta': {'href': 'https://api.moysklad.ru/api/remap/1.2/entity/product/idle'},
        'article': '99-00XX0000', 'code': '2-999', 'name': 'Товар без движения',
        'folder': {'name': 'ArtroPro'}, 'stock': 0.0, 'reserve': 0.0, 'inTransit': 0.0, 'quantity': 0.0,
    },
]

PRODUCT_ROWS = [
    {'id': 'probiotic', 'pathName': 'Товары/GastroPro', 'minimumBalance': 267.0},
    {'id': 'biotin', 'pathName': 'Товары/ExterPro', 'minimumBalance': 133.0},
    {'id': 'gel', 'pathName': 'Товары/ArtroPro', 'minimumBalance': 29.0},
    {'id': 'idle', 'pathName': 'Товары/ArtroPro', 'minimumBalance': 0.0},
    # Чужие ветки: фильтр pathName~Товары на стороне МойСклад ищет вхождение
    # подстроки и приносит ещё и эти две группы вместе с их минимумами
    {'id': 'foreign-marketplace', 'pathName': 'Товары маркетплейсов/Магазин на Ozon', 'minimumBalance': 500.0},
    {'id': 'foreign-outdoor', 'pathName': 'Товары на выезд', 'minimumBalance': 20.0},
]


def fake_ms_get(url, **kwargs):
    """Ответы МойСклад по четырём эндпоинтам, которые дёргает страница."""
    response = MagicMock()
    if '/entity/store' in url:
        response.json.return_value = {'rows': [{'meta': {'href': STORE_HREF}}], 'meta': {'size': 1}}
    elif '/entity/productfolder' in url:
        response.json.return_value = {'rows': [{'name': 'Товары', 'meta': {'href': FOLDER_HREF}}], 'meta': {'size': 1}}
    elif '/entity/product' in url:
        response.json.return_value = {'rows': PRODUCT_ROWS, 'meta': {'size': len(PRODUCT_ROWS)}}
    elif '/report/stock/all' in url:
        response.json.return_value = {'rows': STOCK_ROWS, 'meta': {'size': len(STOCK_ROWS)}}
    else:
        raise AssertionError(f'Неожиданный запрос к МойСклад: {url}')
    return response


class FboStockCalculationTests(SimpleTestCase):
    def setUp(self):
        cache.clear()

    def tearDown(self):
        cache.clear()

    @patch('api.views.fbo_stock.ms_http.get', side_effect=fake_ms_get)
    def test_in_transit_is_not_counted_as_available(self, _mock_get):
        """«Можно взять» = остаток − резерв: плановый выпуск в него не входит."""
        items = {item['article']: item for item in _build_data()['items']}

        probiotic = items['06-03GP0800']
        self.assertEqual(probiotic['quantity'], 219.0)      # не 289, как «Доступно» в МойСклад
        self.assertEqual(probiotic['in_transit'], 70.0)

        self.assertEqual(items['08-02EP0800']['quantity'], 201.0)

    @patch('api.views.fbo_stock.ms_http.get', side_effect=fake_ms_get)
    def test_reserve_above_stock_stays_negative(self, _mock_get):
        """Резерв больше остатка — показываем минус, а не ноль: это повод разобраться."""
        items = {item['article']: item for item in _build_data()['items']}
        self.assertEqual(items['01-14AP0250']['quantity'], -8.0)

    @patch('api.views.fbo_stock.ms_http.get', side_effect=fake_ms_get)
    def test_below_minimum_marks_only_positions_under_their_own_minimum(self, _mock_get):
        items = {item['article']: item for item in _build_data()['items']}
        self.assertTrue(items['06-03GP0800']['below_minimum'])    # 219 < 267
        self.assertFalse(items['08-02EP0800']['below_minimum'])   # 201 > 133
        self.assertFalse(items['99-00XX0000']['below_minimum'])   # минимум не задан

    @patch('api.views.fbo_stock.ms_http.get', side_effect=fake_ms_get)
    def test_minimum_balances_ignore_other_folder_branches(self, _mock_get):
        """Минимумы берём только из группы «Товары», не из соседних веток."""
        minimums = _minimum_balances()

        self.assertEqual(minimums['probiotic'], 267.0)
        # «Товары маркетплейсов» и «Товары на выезд» начинаются с «Товары»,
        # но это другие группы со своими минимумами
        self.assertNotIn('foreign-marketplace', minimums)
        self.assertNotIn('foreign-outdoor', minimums)

    @patch('api.views.fbo_stock.ms_http.get', side_effect=fake_ms_get)
    def test_stock_report_is_grouped_by_product(self, mock_get):
        """Без groupBy отчёт группирует по модификациям, и минимум товара не находится."""
        _build_data()

        report_calls = [call for call in mock_get.call_args_list if '/report/stock/all' in call.args[0]]
        self.assertTrue(report_calls)
        for call in report_calls:
            self.assertEqual(call.kwargs['params']['groupBy'], 'product')

    @patch('api.views.fbo_stock.ms_http.get', side_effect=fake_ms_get)
    def test_positions_without_movement_are_marked_empty(self, _mock_get):
        items = {item['article']: item for item in _build_data()['items']}
        self.assertTrue(items['99-00XX0000']['is_empty'])
        self.assertFalse(items['06-03GP0800']['is_empty'])


class FboStockEndpointTests(TestCase):
    def setUp(self):
        cache.clear()
        self.client = Client()
        self.user = User.objects.create_user(username='fbo-stock-user', password='pw')
        UserPageAccess.objects.create(user=self.user, page_key='fbo-stock')
        self.client.login(username='fbo-stock-user', password='pw')

    def tearDown(self):
        cache.clear()

    @patch('api.views.fbo_stock.ms_http.get', side_effect=fake_ms_get)
    def test_endpoint_returns_items(self, _mock_get):
        response = self.client.get('/api/analysis/fbo-stock/')
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['status'], 'success')
        articles = {item['article'] for item in payload['data']['items']}
        self.assertIn('06-03GP0800', articles)

    @patch('api.views.fbo_stock.ms_http.get', side_effect=fake_ms_get)
    def test_excel_has_four_columns_and_skips_idle_positions(self, _mock_get):
        response = self.client.get('/api/analysis/fbo-stock/export/')
        self.assertEqual(response.status_code, 200)

        sheet = load_workbook(BytesIO(response.content)).active
        # Строка 1 — заголовок отчёта, строка 2 — шапка таблицы
        self.assertEqual(
            [sheet.cell(row=2, column=col).value for col in range(1, 5)],
            ['Артикул', 'Название', 'Количество', 'Ожидание'],
        )
        rows = {sheet.cell(row=r, column=1).value: sheet.cell(row=r, column=3).value
                for r in range(3, sheet.max_row + 1)}
        self.assertEqual(rows['06-03GP0800'], 219.0)
        self.assertNotIn('99-00XX0000', rows)   # позиции без движения в выгрузку не идут

    @patch('api.views.fbo_stock.ms_http.get', side_effect=fake_ms_get)
    def test_access_is_separate_from_fbo_orders_page(self, _mock_get):
        """Соседний префикс /api/analysis/fbo/ не должен «съедать» страницу остатков."""
        self.assertEqual(page_keys_for_path('/api/analysis/fbo-stock/'), {'fbo-stock'})
        self.assertEqual(page_keys_for_path('/api/analysis/fbo/'), {'fbo'})

        other = User.objects.create_user(username='fbo-orders-user', password='pw')
        UserPageAccess.objects.create(user=other, page_key='fbo')
        other_client = Client()
        other_client.login(username='fbo-orders-user', password='pw')
        self.assertEqual(other_client.get('/api/analysis/fbo-stock/').status_code, 403)
