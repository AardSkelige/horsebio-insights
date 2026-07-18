"""
Cash flow business logic services.
Extracted from views/cash_flow.py for separation of concerns.
"""
import requests
from concurrent.futures import ThreadPoolExecutor
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar
from io import BytesIO

import logging

logger = logging.getLogger(__name__)

BASE_URL = 'https://api.moysklad.ru/api/remap/1.2'
TIMEOUT = 30


def get_headers():
    """Получить заголовки для API запросов"""
    return {
        'Authorization': f'Bearer {settings.MOYSKLAD_TOKEN}',
        'Accept': 'application/json;charset=utf-8',
        'Content-Type': 'application/json'
    }


def _fetch_operation_pages(operation, filter_from, filter_to, headers):
    """Постраничная загрузка всех операций одного типа"""
    all_operations = []
    offset = 0
    limit = 1000

    while True:
        url = f"{BASE_URL}/entity/{operation}?filter=moment>={filter_from};moment<={filter_to}&limit={limit}&offset={offset}"

        try:
            response = requests.get(url, headers=headers, timeout=TIMEOUT)
            response.raise_for_status()
            data = response.json()
        except Exception:
            # Частичный набор финансовых операций нельзя использовать для отчёта:
            # вызывающий view преобразует ошибку API в корректный error response.
            logger.exception("Ошибка получения операций %s", operation)
            raise

        rows = data.get('rows', [])

        if not rows:  # Больше нет данных
            break

        all_operations.extend(rows)
        offset += limit

        # Если получили меньше чем limit, значит это последняя страница
        if len(rows) < limit:
            break

    logger.info(f"Итого операций {operation}: {len(all_operations)}")
    return all_operations


def get_operations_data(date_from, date_to):
    """
    Получить данные по операциям (все типы, влияющие на денежный поток)
    С постраничной загрузкой для получения всех операций
    """
    headers = get_headers()
    # Включаем все типы операций, влияющие на денежные средства
    operations = [
        'cashin', 'cashout', 'paymentin', 'paymentout',
        'retaildrawercashin', 'retaildrawercashout'
    ]
    filter_from = date_from.replace('T', ' ').replace('.000', '')
    filter_to = date_to.replace('T', ' ').replace('.000', '')

    # Типы операций независимы — грузим параллельно.
    # МойСклад допускает до 5 одновременных запросов, держим запас.
    with ThreadPoolExecutor(max_workers=4) as pool:
        results = pool.map(
            lambda op: (op, _fetch_operation_pages(op, filter_from, filter_to, headers)),
            operations,
        )

    return dict(results)


def get_expense_items():
    """
    Получить список статей расходов из МойСклад
    """
    headers = get_headers()
    url = f"{BASE_URL}/entity/expenseitem?limit=100"
    expense_items = {}

    try:
        response = requests.get(url, headers=headers, timeout=TIMEOUT)
        if response.status_code == 200:
            data = response.json()
            items = data.get('rows', [])

            for item in items:
                item_id = item.get('id')
                name = item.get('name', 'Без названия')
                expense_items[item_id] = name

        else:
            logger.error(f"Ошибка получения статей расходов: {response.status_code}")

    except Exception as e:
        logger.error(f"Ошибка запроса статей расходов: {e}")

    return expense_items


def get_expense_categories_by_api(operations_data, expense_items_dict):
    """
    Извлечь и сгруппировать расходы по статьям расходов из API МойСклад
    Возвращает также ID статей расходов для генерации ссылок МойСклад
    """
    expense_categories = {}
    excluded_categories = {}  # Исключенные категории
    expense_categories_with_ids = {}  # Категории с ID для ссылок
    excluded_categories_with_ids = {}  # Исключенные категории с ID

    # Анализируем расходные операции
    for operation_type in ['cashout', 'paymentout']:
        if operation_type in operations_data:
            for operation in operations_data[operation_type]:
                sum_value = operation.get('sum', 0)
                amount = sum_value / 100 if sum_value else 0
                name = operation.get('name', 'Без названия')

                # Получаем статью расходов из поля expenseItem
                expense_item = operation.get('expenseItem')
                category = 'Без категории'
                expense_item_id = None

                if expense_item and expense_item.get('meta'):
                    href = expense_item['meta'].get('href', '')
                    if '/entity/expenseitem/' in href:
                        expense_item_id = href.split('/entity/expenseitem/')[-1]
                        category = expense_items_dict.get(expense_item_id, f'Неизвестная статья ({expense_item_id})')

                # Исключаем "Перемещение" из основных расходов
                if category == 'Перемещение':
                    if category in excluded_categories:
                        excluded_categories[category] += amount
                    else:
                        excluded_categories[category] = amount

                    # Сохраняем ID для исключенных категорий
                    if expense_item_id:
                        excluded_categories_with_ids[category] = expense_item_id
                else:
                    if category in expense_categories:
                        expense_categories[category] += amount
                    else:
                        expense_categories[category] = amount

                    # Сохраняем ID для основных категорий
                    if expense_item_id:
                        expense_categories_with_ids[category] = expense_item_id

    return expense_categories, excluded_categories, expense_categories_with_ids, excluded_categories_with_ids


def get_sales_channels():
    """
    Получить список каналов продаж из МойСклад
    """
    headers = get_headers()
    channels_dict = {}

    # Получаем каналы продаж
    url = BASE_URL + '/entity/saleschannel'

    try:
        response = requests.get(url, headers=headers, timeout=TIMEOUT)
        if response.status_code == 200:
            data = response.json()
            channels = data.get('rows', [])

            # Создаем словарь ID -> Название канала
            for channel in channels:
                channel_id = channel.get('id')
                channel_name = channel.get('name', 'Без названия')
                if channel_id:
                    channels_dict[channel_id] = channel_name

        else:
            logger.error(f"Ошибка получения каналов: {response.status_code}")

    except Exception as e:
        logger.error(f"Ошибка запроса каналов: {e}")

    return channels_dict


def get_payments_by_channels(channels_dict, operations_data):
    """
    Получить суммы ПЛАТЕЖЕЙ по каналам в строгом режиме.
    Работает по уже загруженным операциям (get_operations_data) — без повторных
    запросов к МойСклад; заодно охватывает все платежи, а не первую тысячу.
    """
    channel_payments = {}
    # Инициализируем каналы нулевыми значениями
    for channel_id, channel_name in channels_dict.items():
        channel_payments[channel_name] = 0
    channel_payments['Без канала'] = 0

    # Добавляем обработку для неизвестных каналов
    def ensure_channel_exists(channel_name):
        if channel_name not in channel_payments:
            channel_payments[channel_name] = 0

    processed_payments = set()  # Отслеживаем обработанные платежи по ID

    for payment_type in ['paymentin', 'cashin']:
        for payment in operations_data.get(payment_type, []):
            # Проверяем, не обрабатывали ли мы уже этот платеж
            payment_id = payment.get('id')
            if payment_id in processed_payments:
                continue
            processed_payments.add(payment_id)

            sum_value = payment.get('sum', 0)
            amount = sum_value / 100 if sum_value else 0
            channel_found = False

            # Проверяем прямое поле salesChannel у платежа
            sales_channel = payment.get('salesChannel')
            if sales_channel and sales_channel.get('meta'):
                href = sales_channel['meta'].get('href', '')
                if '/entity/saleschannel/' in href:
                    channel_id = href.split('/entity/saleschannel/')[-1]
                    channel_name = channels_dict.get(channel_id, f'Неизвестный канал ({channel_id})')
                    ensure_channel_exists(channel_name)
                    channel_payments[channel_name] += amount
                    channel_found = True

            # Если никак не найден
            if not channel_found:
                channel_payments['Без канала'] += amount

    return channel_payments


def get_income_channels_from_payments(channel_payments, operations_data):
    """
    Объединяем данные о платежах по каналам с операциями платежей
    """
    # Сначала используем реальные данные платежей по каналам
    income_channels = dict(channel_payments)

    # Добавляем прочие доходы из операций (не связанные с продажами)
    total_operations_income = 0
    total_payments_income = sum(amount for amount in channel_payments.values() if amount > 0)

    # Считаем общий доход из операций
    for operation_type in ['cashin', 'paymentin']:
        if operation_type in operations_data:
            for operation in operations_data[operation_type]:
                sum_value = operation.get('sum', 0)
                amount = sum_value / 100 if sum_value else 0
                total_operations_income += amount

    # Если есть разница между общими операциями и платежами - это прочие доходы
    other_income = total_operations_income - total_payments_income
    if other_income > 0:
        income_channels['Прочие доходы'] = other_income

    return income_channels


def get_last_full_month_dates():
    """
    Получить даты начала и конца последнего полного месяца
    """
    today = datetime.now()
    # Первый день предыдущего месяца
    if today.month == 1:
        first_day = datetime(today.year - 1, 12, 1)
    else:
        first_day = datetime(today.year, today.month - 1, 1)

    # Последний день предыдущего месяца
    last_day = datetime(first_day.year, first_day.month, calendar.monthrange(first_day.year, first_day.month)[1])

    date_from = first_day.strftime('%Y-%m-%dT00:00:00.000')
    date_to = last_day.strftime('%Y-%m-%dT23:59:59.000')

    return date_from, date_to


def generate_moysklad_link(date_from, date_to, expense_item_id=None, expense_item_name=None):
    """
    Генерирует ссылку на МойСклад с фильтрами по дате и статье расходов

    Args:
        date_from: Дата начала в формате YYYY-MM-DDTHH:mm:ss.SSS
        date_to: Дата окончания в формате YYYY-MM-DDTHH:mm:ss.SSS
        expense_item_id: ID статьи расходов
        expense_item_name: Название статьи расходов

    Returns:
        URL для МойСклад с настроенными фильтрами
    """
    base_url = "https://online.moysklad.ru/app/#finance"

    # Конвертируем даты из ISO формата в формат МойСклад
    try:
        date_from_dt = datetime.fromisoformat(date_from.replace('T', ' ').replace('.000', ''))
        date_to_dt = datetime.fromisoformat(date_to.replace('T', ' ').replace('.000', ''))

        # Форматируем даты для URL МойСклад: DD.MM.YYYY HH:MM:SS
        formatted_date_from = date_from_dt.strftime('%d.%m.%Y %H:%M:%S')
        formatted_date_to = date_to_dt.strftime('%d.%m.%Y %H:%M:%S')
    except Exception as e:
        logger.error(f"Ошибка форматирования дат: {e}")
        return base_url

    # Создаем параметры фильтрации без автоматического URL encoding
    params = []

    # Фильтр по периоду - НЕ кодируем двоеточия и запятые внутри значения
    period_filter = f"{formatted_date_from},{formatted_date_to},inside_period"
    # Кодируем только пробелы и специальные символы, оставляем : и , как есть
    period_encoded = period_filter.replace(' ', '%20')
    params.append(f"global_periodFilter={period_encoded}")

    # Фильтр по статье расходов (если указан)
    if expense_item_id and expense_item_name:
        # Формат: [ID\\,Название\\,Название\\,ExpenseItem],equals
        expense_filter = f"[{expense_item_id}\\\\,{expense_item_name}\\\\,{expense_item_name}\\\\,ExpenseItem],equals"
        # Кодируем только квадратные скобки и специальные символы, оставляем запятые и слэши
        expense_encoded = expense_filter.replace('[', '%5B').replace(']', '%5D')
        params.append(f"global_expenseItemFilter={expense_encoded}")

    # Собираем финальный URL
    if params:
        return f"{base_url}?{'&'.join(params)}"
    else:
        return base_url


def get_earliest_operation_date():
    """
    Получить дату самой ранней операции в МойСклад
    """
    headers = get_headers()
    operations = ['cashin', 'cashout', 'paymentin', 'paymentout']
    earliest_date = None

    for operation in operations:
        url = f"{BASE_URL}/entity/{operation}?limit=1&order=moment,asc"

        try:
            response = requests.get(url, headers=headers, timeout=TIMEOUT)
            if response.status_code == 200:
                data = response.json()
                rows = data.get('rows', [])
                if rows:
                    moment = rows[0].get('moment')
                    if moment:
                        operation_date = datetime.fromisoformat(moment.replace('T', ' ').replace('.000', ''))
                        if earliest_date is None or operation_date < earliest_date:
                            earliest_date = operation_date
        except Exception as e:
            logger.error(f"Ошибка получения ранней операции {operation}: {e}")

    return earliest_date


def calculate_initial_balance(start_date):
    """
    Рассчитать начальный остаток на определенную дату
    Используя базовую точку - январь 2025
    """
    logger.info(f"Расчет начального остатка на {start_date}:")

    # Базовая точка отсчета - данные из МойСклад за январь 2025
    base_date = '2025-01-01T00:00:00.000'
    base_initial_balance = 8877300.28  # Начальный остаток на 01.01.2025

    # Конвертируем даты для сравнения
    start_datetime = datetime.fromisoformat(start_date.replace('T', ' ').replace('.000', ''))
    base_datetime = datetime.fromisoformat(base_date.replace('T', ' ').replace('.000', ''))

    logger.info(f"  Базовая точка: {base_date} = {base_initial_balance:,.2f} руб")

    # Если запрашиваемая дата совпадает с базовой
    if start_datetime == base_datetime:
        logger.info("  Точное совпадение с базовой точкой")
        return base_initial_balance

    # Если запрашиваемая дата раньше базовой точки
    if start_datetime < base_datetime:
        logger.info(f"  Дата {start_date} раньше базовой точки")
        logger.info("  Невозможно рассчитать без данных до 2025 года")
        return base_initial_balance  # Возвращаем базовое значение

    # Рассчитываем остаток на start_date от базовой точки
    # Получаем все операции от 01.01.2025 до start_date
    operations_data = get_operations_data(base_date, start_date)

    total_income = 0
    total_expense = 0
    operations_count = 0

    for operation_type, operations in operations_data.items():
        if not operations:
            continue

        for operation in operations:
            operation_date = operation.get('moment', '')
            if not operation_date:
                continue

            operation_datetime = datetime.fromisoformat(operation_date.replace('T', ' ').replace('.000', ''))

            # Учитываем операции строго ДО start_date
            if operation_datetime >= start_datetime:
                continue

            sum_value = operation.get('sum', 0)
            amount = sum_value / 100 if sum_value else 0
            operations_count += 1

            if operation_type in ['cashout', 'paymentout', 'retaildrawercashout']:
                total_expense += amount
            elif operation_type in ['cashin', 'paymentin', 'retaildrawercashin']:
                total_income += amount

    # Начальный остаток = базовый остаток + операции с базовой даты до start_date
    initial_balance = base_initial_balance + total_income - total_expense

    logger.info(f"  Операций с {base_date} до {start_date}: {operations_count}")
    logger.info(f"  Доходы с базовой даты: {total_income:,.2f} руб")
    logger.info(f"  Расходы с базовой даты: {total_expense:,.2f} руб")
    logger.info(f"  Начальный остаток: {initial_balance:,.2f} руб")

    return initial_balance


def export_to_excel(income_channels, expense_categories, total_income, total_expense, period_from, period_to, initial_balance, excluded_categories=None):
    """
    Экспорт данных о движении денежных средств в Excel файл
    """
    # Создаем новую рабочую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Движение денежных средств"

    # Стили для форматирования
    header_font = Font(bold=True, size=12)
    subheader_font = Font(bold=True, size=11)

    header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    income_fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
    expense_fill = PatternFill(start_color='FFE8E8', end_color='FFE8E8', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовок отчета
    current_row = 1
    ws.merge_cells(f'A{current_row}:C{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "ОТЧЕТ О ДВИЖЕНИИ ДЕНЕЖНЫХ СРЕДСТВ"
    cell.font = Font(bold=True, size=14)
    cell.fill = header_fill
    cell.border = thin_border
    current_row += 1

    # Период
    ws.merge_cells(f'A{current_row}:C{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = f"Период: {period_from} — {period_to}"
    cell.font = subheader_font
    cell.border = thin_border
    current_row += 2

    # Заголовки колонок
    ws[f'A{current_row}'] = "Статья"
    ws[f'B{current_row}'] = "Сумма, руб"
    ws[f'C{current_row}'] = "Примечание"

    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{current_row}']
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    current_row += 1

    # Начальный остаток (передается как параметр)
    ws[f'A{current_row}'] = "Начальный остаток"
    ws[f'B{current_row}'] = initial_balance
    ws[f'C{current_row}'] = "Остаток на начало периода"

    for col in ['A', 'B', 'C']:
        ws[f'{col}{current_row}'].border = thin_border
    current_row += 1

    # ПРИХОДЫ
    ws[f'A{current_row}'] = "ПРИХОДЫ"
    ws[f'B{current_row}'] = total_income
    ws[f'C{current_row}'] = "Общая сумма поступлений"

    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{current_row}']
        cell.font = subheader_font
        cell.fill = income_fill
        cell.border = thin_border
    current_row += 1

    # Детализация приходов по каналам
    for channel, amount in income_channels.items():
        if amount > 0:
            ws[f'A{current_row}'] = f"  {channel}"
            ws[f'B{current_row}'] = amount
            ws[f'C{current_row}'] = "Канал продаж"

            for col in ['A', 'B', 'C']:
                ws[f'{col}{current_row}'].border = thin_border
            current_row += 1

    # Пустая строка
    current_row += 1

    # РАСХОДЫ
    ws[f'A{current_row}'] = "РАСХОДЫ"
    ws[f'B{current_row}'] = total_expense
    ws[f'C{current_row}'] = "Общая сумма расходов"

    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{current_row}']
        cell.font = subheader_font
        cell.fill = expense_fill
        cell.border = thin_border
    current_row += 1

    # Детализация расходов по статьям
    for category, amount in sorted(expense_categories.items(), key=lambda x: x[1], reverse=True):
        if amount > 0:
            ws[f'A{current_row}'] = f"  {category}"
            ws[f'B{current_row}'] = amount
            ws[f'C{current_row}'] = "Статья расходов"

            for col in ['A', 'B', 'C']:
                ws[f'{col}{current_row}'].border = thin_border
            current_row += 1

    # Пустая строка
    current_row += 1

    # ИСКЛЮЧЕННЫЕ ОПЕРАЦИИ
    if excluded_categories:
        ws[f'A{current_row}'] = "ИСКЛЮЧЕННЫЕ ОПЕРАЦИИ"
        total_excluded = sum(excluded_categories.values())
        ws[f'B{current_row}'] = total_excluded
        ws[f'C{current_row}'] = "Операции не учитываются в расчетах"

        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{current_row}']
            cell.font = subheader_font
            cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            cell.border = thin_border
        current_row += 1

        # Детализация исключенных операций
        for category, amount in excluded_categories.items():
            if amount > 0:
                ws[f'A{current_row}'] = f"  {category}"
                ws[f'B{current_row}'] = amount
                ws[f'C{current_row}'] = "Исключено"

                for col in ['A', 'B', 'C']:
                    cell = ws[f'{col}{current_row}']
                    cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                    cell.border = thin_border
                current_row += 1

        # Пустая строка
        current_row += 1

    # ИТОГО
    net_cash_flow = total_income - total_expense
    profit = total_income - total_expense
    final_balance = initial_balance + net_cash_flow

    ws[f'A{current_row}'] = "Чистый денежный поток"
    ws[f'B{current_row}'] = net_cash_flow
    ws[f'C{current_row}'] = "Приход - Расход"

    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{current_row}']
        cell.font = subheader_font
        cell.fill = header_fill
        cell.border = thin_border
    current_row += 1

    ws[f'A{current_row}'] = "Прибыль"
    ws[f'B{current_row}'] = profit
    ws[f'C{current_row}'] = "Общий приход - Общий расход"

    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{current_row}']
        cell.font = subheader_font
        cell.fill = header_fill
        cell.border = thin_border
    current_row += 1

    ws[f'A{current_row}'] = "Конечный остаток"
    ws[f'B{current_row}'] = final_balance
    ws[f'C{current_row}'] = "Остаток на конец периода"

    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{current_row}']
        cell.font = subheader_font
        cell.fill = header_fill
        cell.border = thin_border

    # Автоматическая ширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError):
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

    # Форматирование чисел
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            if cell.column == 2 and isinstance(cell.value, (int, float)):  # Колонка B (суммы)
                cell.number_format = '# ### ###'

    # Сохраняем в BytesIO для возврата
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer
