"""
Excel export utilities for API views.
Consolidates duplicated Excel generation code from multiple views.
"""
from io import BytesIO
from typing import Any, Callable, Dict, List, Optional, Union

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


# Default styles for Excel exports
HEADER_FONT = Font(bold=True)
HEADER_FONT_WHITE = Font(bold=True, color='FFFFFF')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FILL_GRAY = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


class ExcelReportBuilder:
    """
    Builder class for creating Excel reports with consistent styling.

    Example:
        >>> builder = ExcelReportBuilder("Report Title")
        >>> builder.add_headers(['Name', 'Value', 'Total'])
        >>> builder.add_row(['Product A', 100, 1000])
        >>> builder.add_row(['Product B', 200, 2000])
        >>> builder.set_column_widths([40, 15, 20])
        >>> response = builder.to_http_response('report.xlsx')
    """

    def __init__(self, title: str = "Sheet1"):
        """Initialize the Excel builder with a workbook and active sheet."""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = title
        self.current_row = 1
        self._headers_added = False

    def add_headers(
        self,
        headers: List[str],
        use_fill: bool = True,
        fill_color: str = '4472C4',
        white_text: bool = True
    ) -> 'ExcelReportBuilder':
        """
        Add header row with styling.

        Args:
            headers: List of header titles
            use_fill: Whether to use background fill
            fill_color: Hex color for background
            white_text: Whether to use white text

        Returns:
            Self for chaining
        """
        font = HEADER_FONT_WHITE if white_text else HEADER_FONT
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid') if use_fill else None

        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col, value=header)
            cell.font = font
            cell.alignment = HEADER_ALIGNMENT
            if fill:
                cell.fill = fill
            cell.border = THIN_BORDER

        self.current_row += 1
        self._headers_added = True
        return self

    def add_row(
        self,
        values: List[Any],
        number_formats: Optional[Dict[int, str]] = None,
        font_colors: Optional[Dict[int, str]] = None
    ) -> 'ExcelReportBuilder':
        """
        Add a data row.

        Args:
            values: List of cell values
            number_formats: Dict mapping column index to number format (e.g., {2: '#,##0.00'})
            font_colors: Dict mapping column index to font color (e.g., {2: 'FF0000'})

        Returns:
            Self for chaining
        """
        for col, value in enumerate(values, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col, value=value)
            cell.border = THIN_BORDER

            if number_formats and col in number_formats:
                cell.number_format = number_formats[col]

            if font_colors and col in font_colors:
                cell.font = Font(color=font_colors[col])

        self.current_row += 1
        return self

    def add_rows_from_data(
        self,
        data: List[Dict[str, Any]],
        columns: List[str],
        formatters: Optional[Dict[str, Callable]] = None,
        number_formats: Optional[Dict[int, str]] = None
    ) -> 'ExcelReportBuilder':
        """
        Add multiple rows from a list of dictionaries.

        Args:
            data: List of dictionaries containing row data
            columns: List of dictionary keys to extract in order
            formatters: Dict mapping column name to formatter function
            number_formats: Dict mapping column index to number format

        Returns:
            Self for chaining
        """
        for item in data:
            values = []
            for col_name in columns:
                value = item.get(col_name, '')
                if formatters and col_name in formatters:
                    value = formatters[col_name](value)
                values.append(value)
            self.add_row(values, number_formats=number_formats)

        return self

    def set_column_widths(self, widths: List[int]) -> 'ExcelReportBuilder':
        """
        Set column widths.

        Args:
            widths: List of width values for each column

        Returns:
            Self for chaining
        """
        for col, width in enumerate(widths, 1):
            col_letter = self.worksheet.cell(row=1, column=col).column_letter
            self.worksheet.column_dimensions[col_letter].width = width

        return self

    def add_title_row(self, title: str, merge_cols: int = 1) -> 'ExcelReportBuilder':
        """
        Add a title row that spans multiple columns.

        Args:
            title: Title text
            merge_cols: Number of columns to merge

        Returns:
            Self for chaining
        """
        cell = self.worksheet.cell(row=self.current_row, column=1, value=title)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

        if merge_cols > 1:
            self.worksheet.merge_cells(
                start_row=self.current_row,
                start_column=1,
                end_row=self.current_row,
                end_column=merge_cols
            )

        self.current_row += 1
        return self

    def create_sheet(self, title: str) -> 'ExcelReportBuilder':
        """
        Create a new sheet and make it active.

        Args:
            title: Sheet title

        Returns:
            Self for chaining
        """
        self.worksheet = self.workbook.create_sheet(title)
        self.current_row = 1
        self._headers_added = False
        return self

    def to_bytes(self) -> BytesIO:
        """
        Get workbook as BytesIO buffer.

        Returns:
            BytesIO buffer containing the Excel file
        """
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def to_http_response(self, filename: str) -> HttpResponse:
        """
        Create HTTP response for file download.

        Args:
            filename: Filename for the download

        Returns:
            HttpResponse with Excel content
        """
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        self.workbook.save(response)
        return response


def create_excel_response(
    headers: List[str],
    data: List[List[Any]],
    filename: str,
    sheet_title: str = "Data",
    column_widths: Optional[List[int]] = None
) -> HttpResponse:
    """
    Create a simple Excel file and return as HTTP response.

    Args:
        headers: List of column headers
        data: List of rows, each row is a list of values
        filename: Filename for the download
        sheet_title: Title for the worksheet
        column_widths: Optional list of column widths

    Returns:
        HttpResponse with Excel content

    Example:
        >>> headers = ['Name', 'Price', 'Quantity']
        >>> data = [['Product A', 100, 10], ['Product B', 200, 20]]
        >>> response = create_excel_response(headers, data, 'products.xlsx')
    """
    builder = ExcelReportBuilder(sheet_title)
    builder.add_headers(headers)

    for row in data:
        builder.add_row(row)

    if column_widths:
        builder.set_column_widths(column_widths)
    else:
        # Auto-calculate widths based on header length
        widths = [max(len(str(h)) + 2, 12) for h in headers]
        builder.set_column_widths(widths)

    return builder.to_http_response(filename)


def auto_adjust_column_widths(worksheet: Worksheet, max_width: int = 50) -> None:
    """
    Auto-adjust column widths based on content.

    Args:
        worksheet: Worksheet to adjust
        max_width: Maximum column width
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column_letter

        for cell in column_cells:
            try:
                cell_length = len(str(cell.value or ''))
                if cell_length > max_length:
                    max_length = cell_length
            except (TypeError, AttributeError):
                pass

        adjusted_width = min(max_length + 2, max_width)
        worksheet.column_dimensions[column].width = adjusted_width
