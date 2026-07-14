from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, F, Sum, Count, Avg, Min, Max, FloatField, ExpressionWrapper
from django.db.models.fields import DecimalField
from rest_framework.decorators import api_view
from django.db.models.functions import (
    Coalesce
)
from datetime import datetime
from django.utils import timezone

from core.models import (
    Product,
    ShipmentItem,
    RawMaterialUsage
)
from api.exceptions import NotFoundError, DataProcessingError

import logging
logger = logging.getLogger(__name__)

@api_view(['GET'])
def product_data(request):
    """API endpoint для получения данных по товарам с фильтрацией"""
    try:
        # Получаем параметры фильтрации
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('pageSize', 10))
        search = request.GET.get('search', '').strip()
        subgroup = request.GET.get('subgroup', '').strip()
        start_date = request.GET.get('startDate')
        end_date = request.GET.get('endDate')
        sort_field = request.GET.get('sortField', 'total_quantity')
        sort_order = request.GET.get('sortOrder', 'desc')

        # Базовый QuerySet для товаров с исключением пустых подгрупп
        products_query = Product.objects.filter(
            group='Товары'
        ).exclude(
            subgroup__isnull=True
        )

        # Улучшенный поиск по нескольким словам
        if search:
            search_words = search.split()
            search_query = Q()
            for word in search_words:
                word_query = (
                    Q(name__iregex=f'(?i){word}') |
                    Q(article__iregex=f'(?i){word}')
                )
                search_query &= word_query
            products_query = products_query.filter(search_query)

        if subgroup:
            products_query = products_query.filter(subgroup=subgroup)

        # Получаем связанные отгрузки с учетом дат
        shipments_query = ShipmentItem.objects.all()
        
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
                start_date = timezone.make_aware(start_date, timezone.get_current_timezone())
                shipments_query = shipments_query.filter(
                    shipment__date__gte=start_date
                )
            except ValueError:
                pass

        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d')
                # Make end_date inclusive by adding 23:59:59
                end_date = end_date.replace(hour=23, minute=59, second=59)
                end_date = timezone.make_aware(end_date, timezone.get_current_timezone())
                shipments_query = shipments_query.filter(
                    shipment__date__lte=end_date
                )
            except ValueError:
                pass

        # Аггрегируем данные по товарам
        products_data = products_query.annotate(
            total_quantity=Coalesce(
                Sum('shipmentitem__quantity', 
                    filter=Q(shipmentitem__in=shipments_query)),
                0,
                output_field=DecimalField()
            ),
            average_price=Coalesce(
                Avg('shipmentitem__price',
                    filter=Q(shipmentitem__in=shipments_query)),
                0,
                output_field=DecimalField()
            ),
            total_sum=Coalesce(
                Sum(F('shipmentitem__quantity') * F('shipmentitem__price'),
                    filter=Q(shipmentitem__in=shipments_query)),
                0,
                output_field=DecimalField()
            ),
            shipments_count=Count(
                'shipmentitem__shipment',
                filter=Q(shipmentitem__in=shipments_query),
                distinct=True
            )
        )

        # Фильтруем товары с нулевыми показателями, если указан период
        if start_date or end_date:
            products_data = products_data.filter(
                Q(total_quantity__gt=0) | 
                Q(total_sum__gt=0) |
                Q(shipments_count__gt=0)
            )

        # Расчет топ-3 по количеству
        top_quantity = products_data.order_by('-total_quantity')[:3]
        top_quantity_data = [{
            'name': product.name,
            'quantity': float(product.total_quantity),
            'shipments_count': product.shipments_count
        } for product in top_quantity]

        # Расчет топ-3 по выручке
        top_revenue = products_data.order_by('-total_sum')[:3]
        top_revenue_data = [{
            'name': product.name,
            'revenue': float(product.total_sum),
            'price_per_unit': float(product.total_sum / product.total_quantity if product.total_quantity else 0)
        } for product in top_revenue]

        # Расчет топ-3 по среднему количеству
        top_average = products_data.filter(
            shipments_count__gt=0
        ).annotate(
            avg_quantity=ExpressionWrapper(
                F('total_quantity') / F('shipments_count'),
                output_field=FloatField()
            )
        ).order_by('-avg_quantity')[:3]

        top_average_data = [{
            'name': product.name,
            'average_quantity': f"{float(product.avg_quantity):.1f} шт. в среднем за отгрузку"
        } for product in top_average]

        # Статистика
        stats = {
            'total_products': products_data.count(),
            'total_shipments': shipments_query.values('shipment').distinct().count(),
            'top_by_quantity': top_quantity_data,
            'top_by_revenue': top_revenue_data,
            'top_by_average_quantity': top_average_data
        }

        # Применяем сортировку
        if sort_field in ['name', 'subgroup']:
            sort_prefix = '-' if sort_order == 'desc' else ''
            products_data = products_data.order_by(f'{sort_prefix}{sort_field}')
        else:
            sort_mapping = {
                'quantity': 'total_quantity',
                'average_price': 'average_price',
                'total_sum': 'total_sum',
                'shipments_count': 'shipments_count'
            }
            if sort_field in sort_mapping:
                sort_prefix = '-' if sort_order == 'desc' else ''
                products_data = products_data.order_by(f'{sort_prefix}{sort_mapping[sort_field]}')

        # Пагинация
        paginator = Paginator(products_data, page_size)
        page_data = paginator.get_page(page)

        # Форматируем данные для ответа
        products_list = []
        for product in page_data:
            products_list.append({
                'id': product.id,
                'name': product.name,
                'group': product.group,
                'subgroup': product.subgroup,
                'quantity': float(product.total_quantity),
                'average_price': float(product.average_price),
                'total_sum': float(product.total_sum),
                'shipments_count': product.shipments_count
            })

        return JsonResponse({
            'status': 'success',
            'data': {
                'products': products_list,
                'total': paginator.count,
                'stats': stats,
                'available_subgroups': list(products_query.values_list(
                    'subgroup', flat=True
                ).distinct().order_by('subgroup'))
            }
        })

    except Exception as e:
        logger.error(f"Error in product_data: {str(e)}", exc_info=True)
        raise DataProcessingError("Ошибка получения данных по товарам")

@api_view(['GET'])
def product_details(request, product_id):
    """API endpoint для получения детальной информации о товаре"""
    try:
        # Получаем базовую информацию о товаре
        product = Product.objects.get(id=product_id)
        
        # Получаем параметры фильтрации по датам
        start_date = request.GET.get('startDate')
        end_date = request.GET.get('endDate')

        # Базовый QuerySet для отгрузок
        shipments_query = ShipmentItem.objects.filter(
            product=product
        ).select_related('shipment')

        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
                start_date = timezone.make_aware(start_date, timezone.get_current_timezone())
                shipments_query = shipments_query.filter(
                    shipment__date__gte=start_date
                )
            except ValueError:
                pass

        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d')
                # Make end_date inclusive by adding 23:59:59
                end_date = end_date.replace(hour=23, minute=59, second=59)
                end_date = timezone.make_aware(end_date, timezone.get_current_timezone())
                shipments_query = shipments_query.filter(
                    shipment__date__lte=end_date
                )
            except ValueError:
                pass

        # Получаем общую статистику
        stats = {}
        stats_data = shipments_query.aggregate(
            total_shipments=Count('shipment', distinct=True),
            total_quantity=Sum('quantity'),
            min_price=Min('price'),
            max_price=Max('price')
        )
        # Отдельно считаем total_revenue
        total_revenue = sum(item.quantity * item.price for item in shipments_query)
        stats.update(stats_data)

        # Получаем используемые материалы
        materials = RawMaterialUsage.objects.filter(
            shipment_item__in=shipments_query
        ).values(
            'raw_material__name',
            'raw_material__uom_name'
        ).annotate(
            total_quantity=Sum('quantity')
        ).order_by('raw_material__name')

        # Получаем историю отгрузок
        shipments_history = list(shipments_query.order_by('-shipment__date')[:1000])
        
        # Получаем динамику по месяцам
        monthly_data = {}
        for item in shipments_query:
            month = item.shipment.date.replace(day=1)
            if month not in monthly_data:
                monthly_data[month] = {'quantity': 0, 'revenue': 0}
            monthly_data[month]['quantity'] += item.quantity
            monthly_data[month]['revenue'] += item.quantity * item.price

        monthly_dynamics = [
            {
                'month': month,
                'quantity': data['quantity'],
                'revenue': data['revenue']
            }
            for month, data in sorted(monthly_data.items())
        ]

        response_data = {
            'product': {
                'id': product.id,
                'name': product.name,
                'group': product.group,
                'subgroup': product.subgroup
            },
            'statistics': {
                'total_shipments': stats['total_shipments'] or 0,
                'total_quantity': float(stats['total_quantity'] or 0),
                'total_revenue': float(total_revenue),
                'average_quantity': (
                    float(stats['total_quantity'] or 0) / stats['total_shipments'] 
                    if stats['total_shipments'] 
                    else 0
                ),
                'price_range': {
                    'min': float(stats['min_price'] or 0),
                    'max': float(stats['max_price'] or 0)
                }
            },
            'materials': [{
                'name': material['raw_material__name'],
                'quantity': float(material['total_quantity']),
                'unit': material['raw_material__uom_name']
            } for material in materials],
            'shipments_history': [{
                'number': item.shipment.number,
                'date': timezone.localtime(item.shipment.date).isoformat(),
                'quantity': float(item.quantity),
                'price': float(item.price),
                'total': float(item.quantity * item.price)
            } for item in shipments_history],
            'monthly_dynamics': [{
                'month': month.isoformat(),
                'quantity': float(data['quantity']),
                'revenue': float(data['revenue'])
            } for month, data in sorted(monthly_data.items())]
        }

        return JsonResponse({
            'status': 'success',
            'data': response_data
        })

    except Product.DoesNotExist:
        raise NotFoundError("Товар не найден")
    except Exception as e:
        logger.error(f"Error in product_details: {str(e)}", exc_info=True)
        raise DataProcessingError("Ошибка получения детальной информации о товаре")


# Добавим новые импорты
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime

@api_view(['GET'])
def export_products_excel(request):
    """API endpoint для экспорта товаров в Excel"""
    try:
        # Получаем те же параметры фильтрации, что и в product_data
        search = request.GET.get('search', '').strip()
        subgroup = request.GET.get('subgroup', '').strip()
        start_date = request.GET.get('startDate')
        end_date = request.GET.get('endDate')

        # Базовый QuerySet для товаров
        products_query = Product.objects.filter(
            group='Товары'
        ).exclude(
            subgroup__isnull=True
        )

        # Применяем те же фильтры, что и в product_data
        if search:
            search_words = search.split()
            search_query = Q()
            for word in search_words:
                word_query = (
                    Q(name__iregex=f'(?i){word}') |
                    Q(article__iregex=f'(?i){word}')
                )
                search_query &= word_query
            products_query = products_query.filter(search_query)

        if subgroup:
            products_query = products_query.filter(subgroup=subgroup)

        # Получаем связанные отгрузки с учетом дат
        shipments_query = ShipmentItem.objects.all()
        
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
                start_date = timezone.make_aware(start_date, timezone.get_current_timezone())
                shipments_query = shipments_query.filter(
                    shipment__date__gte=start_date
                )
            except ValueError:
                pass

        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d')
                # Make end_date inclusive by adding 23:59:59
                end_date = end_date.replace(hour=23, minute=59, second=59)
                end_date = timezone.make_aware(end_date, timezone.get_current_timezone())
                shipments_query = shipments_query.filter(
                    shipment__date__lte=end_date
                )
            except ValueError:
                pass

        # Аггрегируем данные по товарам
        products_data = products_query.annotate(
            total_quantity=Coalesce(
                Sum('shipmentitem__quantity', 
                    filter=Q(shipmentitem__in=shipments_query)),
                0,
                output_field=DecimalField()
            ),
            average_price=Coalesce(
                Avg('shipmentitem__price',
                    filter=Q(shipmentitem__in=shipments_query)),
                0,
                output_field=DecimalField()
            ),
            total_sum=Coalesce(
                Sum(F('shipmentitem__quantity') * F('shipmentitem__price'),
                    filter=Q(shipmentitem__in=shipments_query)),
                0,
                output_field=DecimalField()
            ),
            shipments_count=Count(
                'shipmentitem__shipment',
                filter=Q(shipmentitem__in=shipments_query),
                distinct=True
            )
        )

        # Создаем новый Excel-файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Товары"

        # Заголовки
        headers = [
            'Товар', 
            'Подгруппа', 
            'Количество', 
            'Средняя цена', 
            'Сумма продаж',
            'Количество отгрузок'
        ]

        # Стили для заголовков
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')

        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment

        # Записываем данные
        for row, product in enumerate(products_data, 2):
            ws.cell(row=row, column=1, value=product.name)
            ws.cell(row=row, column=2, value=product.subgroup)
            ws.cell(row=row, column=3, value=float(product.total_quantity))
            ws.cell(row=row, column=4, value=float(product.average_price))
            ws.cell(row=row, column=5, value=float(product.total_sum))
            ws.cell(row=row, column=6, value=product.shipments_count)

        # Автоматическая ширина колонок
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Создаем HTTP-ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=products_export.xlsx'

        # Сохраняем файл
        wb.save(response)

        return response

    except Exception as e:
        logger.error(f"Error in export_products_excel: {str(e)}", exc_info=True)
        raise DataProcessingError("Ошибка экспорта товаров в Excel")