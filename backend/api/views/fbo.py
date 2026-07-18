from django.http import HttpResponse, JsonResponse
from rest_framework.decorators import api_view
from datetime import datetime, timedelta
import logging
import requests
from django.utils import timezone
from sync.moysklad import MoySkladAPIClient
from core.models import ShipmentItem
from django.conf import settings
from django.db.models import Sum, Value, FloatField
from django.db.models.functions import Coalesce
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from api.exceptions import ExternalServiceError, DataProcessingError

logger = logging.getLogger(__name__)

def get_stock_info(client, assortments):
    """Получение информации об остатках для списка товаров и модификаций"""
    try:
        if not assortments:
            return {}
            
        stock_dict = {}
        batch_size = 10
        
        for i in range(0, len(assortments), batch_size):
            batch = list(assortments)[i:i + batch_size]
            filter_parts = []
            
            for href in batch:
                href = href.split('?')[0]
                if '/variant/' in href:
                    filter_parts.append(f"variant={href}")
                else:
                    filter_parts.append(f"product={href}")
            
            if not filter_parts:
                continue
                
            filter_query = ";".join(filter_parts)
            
            url = f"{client.BASE_URL}/report/stock/all"
            params = {
                "filter": filter_query,
                "groupBy": "product",
                "stockMode": "all"
            }
            
            response = requests.get(url, headers=client.headers, params=params, timeout=30)
            response.raise_for_status()
            data = response.json()

            for item in data.get('rows', []):
                if 'meta' in item and 'href' in item['meta']:
                    meta_href = item['meta']['href'].split('?')[0]
                    stock_data = {
                        'stock': float(item.get('stock', 0)),
                        'quantity': float(item.get('quantity', 0))
                    }
                    stock_dict[meta_href] = stock_data
        
        return stock_dict
        
    except Exception as e:
        return {}

def get_fbo_state_href(client):
    """Href статуса «FBO» из метаданных заказов покупателей"""
    url = f"{client.BASE_URL}/entity/customerorder/metadata"
    response = requests.get(url, headers=client.headers, timeout=30)
    response.raise_for_status()
    for state in response.json().get('states', []):
        if state.get('name') == 'FBO':
            return state['meta']['href']
    return None

@api_view(['GET'])
def get_fbo_analysis(request):
   """API endpoint для анализа FBO заказов"""
   try:
       client = MoySkladAPIClient(settings.MOYSKLAD_TOKEN)

       end_date = timezone.localtime(timezone.now())
       start_date = end_date - timedelta(days=30)
       today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)

       url = f"{client.BASE_URL}/entity/customerorder"
       moment_filter = (
           f"moment>={start_date.strftime('%Y-%m-%d %H:%M:%S')};"
           f"moment<={end_date.strftime('%Y-%m-%d %H:%M:%S')}"
       )

       # Общее число заказов за период — лёгкий запрос без expand, только meta.size
       response = requests.get(url, headers=client.headers,
                               params={"filter": moment_filter, "limit": 1},
                               timeout=30)
       response.raise_for_status()
       total_orders = response.json().get('meta', {}).get('size', 0)

       # FBO-заказы фильтруем на стороне МойСклад (state + deliveryPlannedMoment):
       # вместо всех ~1000 заказов с expand приходит только пара десятков
       state_href = get_fbo_state_href(client)
       if not state_href:
           raise DataProcessingError("Статус «FBO» не найден в МойСклад")

       fbo_filter = (
           f"{moment_filter};state={state_href};"
           f"deliveryPlannedMoment>={today_start.strftime('%Y-%m-%d %H:%M:%S')}"
       )

       candidates = []
       offset = 0
       limit = 100
       while True:
           params = {
               "filter": fbo_filter,
               "limit": limit,
               "offset": offset,
               "expand": "state,agent,positions,positions.assortment,organization,store"
           }
           response = requests.get(url, headers=client.headers, params=params, timeout=30)
           response.raise_for_status()
           rows = response.json().get('rows', [])

           if not rows:
               break

           candidates.extend(rows)

           if len(rows) < limit:
               break

           offset += limit

       # shippedSum не фильтруется в API — оставляем проверку здесь
       fbo_orders = [
           order for order in candidates
           if float(order.get('shippedSum', 0)) == 0 and order.get('deliveryPlannedMoment')
       ]

       products_data = []
       product_hrefs = set()
       
       for order in fbo_orders:
           positions = order.get('positions', {}).get('rows', [])
           for pos in positions:
               assortment = pos.get('assortment', {})
               if assortment.get('meta', {}).get('href'):
                   product_href = assortment['meta']['href'].split('?')[0]
                   product_hrefs.add(product_href)
                   
               product_name = assortment.get('name')
               quantity = float(pos.get('quantity', 0))
               
               product = next(
                   (p for p in products_data if p['name'] == product_name), 
                   None
               )
               
               if product is None:
                   product = {
                       'name': product_name,
                       'fbo_quantity': 0,
                       'href': product_href
                   }
                   products_data.append(product)
               
               product['fbo_quantity'] += quantity

       stock_info = get_stock_info(client, product_hrefs)

       sales_data = dict(
           ShipmentItem.objects
           .filter(
               shipment__date__gte=start_date,
               shipment__date__lte=end_date,
           )
           .values('product__name')
           .annotate(
               total_quantity=Coalesce(
                   Sum('quantity'), 
                   Value(0), 
                   output_field=FloatField()
               )
           )
           .values_list('product__name', 'total_quantity')
       )

       for product in products_data:
           stock_data = stock_info.get(product.get('href', ''), {})
           stock_quantity = float(stock_data.get('stock', 0))
           sales_quantity = float(sales_data.get(product['name'], 0))
           
           product.update({
               'stock': stock_quantity,
               'sales_30_days': sales_quantity,
               'production_needed': product['fbo_quantity'] + sales_quantity - stock_quantity
           })

       products_data.sort(key=lambda x: x['fbo_quantity'], reverse=True)

       orders_data = []
       for order in fbo_orders:
           delivery_date = timezone.make_aware(
               datetime.fromisoformat(
                   order['deliveryPlannedMoment'].replace('Z', '')
               )
           )
           created_date = timezone.make_aware(
               datetime.fromisoformat(
                   order['moment'].replace('Z', '')
               )
           )
           
           orders_data.append({
               'id': order['id'],
               'name': order['name'],
               'created_date': created_date.isoformat(),
               'delivery_date': delivery_date.isoformat(),
               'status': order.get('state', {}).get('name', ''),
               'contractor': order.get('agent', {}).get('name', '')
           })

       response_data = {
           'statistics': {
               'total_orders': total_orders,
               'fbo_orders': len(fbo_orders),
               'no_shipment_orders': len(fbo_orders),
               'start_date': start_date.isoformat(),
               'end_date': end_date.isoformat(),
               'last_update': timezone.now().isoformat()
           },
           'products': products_data,
           'orders': orders_data
       }

       return JsonResponse(response_data)

   except requests.RequestException as e:
       logger.error(f"External API error in get_fbo_analysis: {str(e)}", exc_info=True)
       raise ExternalServiceError("Ошибка подключения к МойСклад API")
   except Exception as e:
       logger.error(f"Error in get_fbo_analysis: {str(e)}", exc_info=True)
       raise DataProcessingError("Ошибка анализа FBO заказов")

@api_view(['GET'])
def export_fbo_excel(request):
    """API endpoint для экспорта FBO анализа в Excel"""
    try:
        response = requests.get(request.build_absolute_uri('/api/analysis/fbo/'), timeout=30)
        data = response.json()
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Создаем лист для товаров
        ws_products = wb.create_sheet("Товары")

        # Добавляем дату отчета
        current_date = timezone.now().strftime('%d.%m.%Y %H:%M')
        ws_products.cell(row=1, column=1, value=f"Отчет сформирован: {current_date}")
        ws_products.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        
        # Стили
        header_style = {
            'font': Font(bold=True),
            'fill': PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid"),
            'alignment': Alignment(horizontal='center', vertical='center')
        }
        date_cell = ws_products.cell(row=1, column=1)
        date_cell.font = Font(bold=True)
        date_cell.alignment = Alignment(horizontal='center')
        
        # Заголовки для товаров
        product_headers = [
            'Наименование',
            'FBO заказ',
            'Остаток',
            'Продажи за 30 дней',
            'Нужно произвести'
        ]
        
        # Применяем стили к заголовкам товаров (сдвигаем на 1 строку вниз из-за даты)
        for col, header in enumerate(product_headers, 1):
            cell = ws_products.cell(row=2, column=col, value=header)
            cell.font = header_style['font']
            cell.fill = header_style['fill']
            cell.alignment = header_style['alignment']

        # Устанавливаем ширину колонок
        ws_products.column_dimensions['A'].width = 60
        ws_products.column_dimensions['B'].width = 15
        ws_products.column_dimensions['C'].width = 15
        ws_products.column_dimensions['D'].width = 20
        ws_products.column_dimensions['E'].width = 20

        # Заполняем данные товаров (сдвигаем на 1 строку вниз из-за даты)
        for row, product in enumerate(data['products'], 3):
            ws_products.cell(row=row, column=1, value=product['name'])
            ws_products.cell(row=row, column=2, value=product['fbo_quantity'])
            ws_products.cell(row=row, column=3, value=product['stock'])
            ws_products.cell(row=row, column=4, value=product['sales_30_days'])
            
            production_cell = ws_products.cell(row=row, column=5, value=product['production_needed'])
            if product['production_needed'] > 0:
                production_cell.font = Font(color="FF0000")
            else:
                production_cell.font = Font(color="008000")

        # Создаем лист для заказов
        ws_orders = wb.create_sheet("Заказы")
        
        # Добавляем дату на лист заказов
        ws_orders.cell(row=1, column=1, value=f"Отчет сформирован: {current_date}")
        ws_orders.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        date_cell_orders = ws_orders.cell(row=1, column=1)
        date_cell_orders.font = Font(bold=True)
        date_cell_orders.alignment = Alignment(horizontal='center')
        
        # Заголовки для заказов
        order_headers = [
            'Номер заказа',
            'Дата создания',
            'Плановая дата отгрузки',
            'Статус',
            'Контрагент'
        ]
        
        # Применяем стили к заголовкам заказов
        for col, header in enumerate(order_headers, 1):
            cell = ws_orders.cell(row=2, column=col, value=header)
            cell.font = header_style['font']
            cell.fill = header_style['fill']
            cell.alignment = header_style['alignment']

        # Устанавливаем ширину колонок для заказов
        ws_orders.column_dimensions['A'].width = 20
        ws_orders.column_dimensions['B'].width = 20
        ws_orders.column_dimensions['C'].width = 20
        ws_orders.column_dimensions['D'].width = 15
        ws_orders.column_dimensions['E'].width = 40

        # Заполняем данные заказов
        for row, order in enumerate(data['orders'], 3):
            ws_orders.cell(row=row, column=1, value=order['name'])
            ws_orders.cell(row=row, column=2, value=datetime.fromisoformat(order['created_date']).strftime('%d.%m.%Y'))
            ws_orders.cell(row=row, column=3, value=datetime.fromisoformat(order['delivery_date']).strftime('%d.%m.%Y'))
            ws_orders.cell(row=row, column=4, value=order['status'])
            ws_orders.cell(row=row, column=5, value=order['contractor'])

        # Создаем HTTP-ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        current_date = timezone.now().strftime('%Y%m%d_%H%M')
        response['Content-Disposition'] = f'attachment; filename=fbo_analysis_{current_date}.xlsx'
        
        wb.save(response)
        return response

    except Exception as e:
        logger.error(f"Error in export_fbo_excel: {str(e)}", exc_info=True)
        raise DataProcessingError("Ошибка экспорта FBO анализа")