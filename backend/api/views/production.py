"""
API для расчёта компонентов производства.

Endpoints:
- POST /api/production/calculate/ - расчёт компонентов из Excel файла
- POST /api/production/calculate-json/ - расчёт компонентов из JSON
- POST /api/production/calculate-single/ - расчёт для одного товара
- POST /api/production/export/ - экспорт результатов в Excel
- GET /api/production/products/search/ - поиск товаров с техкартами
"""
import json
from io import BytesIO
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from api.services.component_calculator import ComponentCalculator


@require_http_methods(["POST"])
def calculate_components_excel(request):
    """
    Расчёт компонентов из Excel файла.

    Request:
        POST multipart/form-data
        - file: Excel файл (.xlsx, .xls)

    Response:
        {
            "success": bool,
            "products_found": int,
            "products_not_found": int,
            "products_without_processing_plan": int,
            "errors": [str],
            "products": [
                {
                    "article": str,
                    "name": str,
                    "quantity": int,
                    "found": bool,
                    "has_processing_plan": bool,
                    "processing_plan_name": str | null,
                    "components": [
                        {
                            "material_id": int,
                            "material_name": str,
                            "uom": str,
                            "quantity_per_unit": float,
                            "total_quantity": float
                        }
                    ],
                    "error": str | null
                }
            ],
            "components_summary": [
                {
                    "material_id": int,
                    "material_name": str,
                    "uom": str,
                    "total_quantity": float
                }
            ]
        }
    """
    if 'file' not in request.FILES:
        return JsonResponse({
            'success': False,
            'error': 'Файл не загружен. Используйте поле "file" для загрузки Excel файла.'
        }, status=400)

    file = request.FILES['file']

    # Проверяем расширение
    if not file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({
            'success': False,
            'error': 'Неверный формат файла. Ожидается .xlsx или .xls'
        }, status=400)

    try:
        calculator = ComponentCalculator()
        file_content = file.read()
        result = calculator.calculate_from_excel(file_content)
        return JsonResponse(result)
    except ValueError as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Ошибка обработки файла: {str(e)}'
        }, status=500)


@require_http_methods(["POST"])
def calculate_components_json(request):
    """
    Расчёт компонентов из JSON.

    Request:
        POST application/json
        {
            "items": [
                {"article": str, "name": str (optional), "quantity": int}
            ]
        }

    Response:
        Такой же как calculate_components_excel
    """
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Неверный формат JSON'
        }, status=400)

    items = data.get('items', [])
    if not items:
        return JsonResponse({
            'success': False,
            'error': 'Список товаров пуст'
        }, status=400)

    # Валидация
    for i, item in enumerate(items):
        if 'article' not in item:
            return JsonResponse({
                'success': False,
                'error': f'Элемент {i}: отсутствует поле "article"'
            }, status=400)
        if 'quantity' not in item:
            return JsonResponse({
                'success': False,
                'error': f'Элемент {i}: отсутствует поле "quantity"'
            }, status=400)
        try:
            item['quantity'] = int(item['quantity'])
        except (ValueError, TypeError):
            return JsonResponse({
                'success': False,
                'error': f'Элемент {i}: "quantity" должно быть числом'
            }, status=400)

    try:
        calculator = ComponentCalculator()
        result = calculator.calculate(items)
        return JsonResponse(result)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Ошибка расчёта: {str(e)}'
        }, status=500)


@require_http_methods(["POST"])
def calculate_single_product(request):
    """
    Расчёт компонентов для одного товара.

    Request:
        POST application/json
        {
            "article": str,
            "quantity": int
        }

    Response:
        {
            "article": str,
            "name": str,
            "quantity": int,
            "found": bool,
            "has_processing_plan": bool,
            "processing_plan_name": str | null,
            "components": [...],
            "error": str | null
        }
    """
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Неверный формат JSON'
        }, status=400)

    article = data.get('article')
    quantity = data.get('quantity')

    if not article:
        return JsonResponse({
            'success': False,
            'error': 'Отсутствует поле "article"'
        }, status=400)

    try:
        quantity = int(quantity or 1)
    except (ValueError, TypeError):
        return JsonResponse({
            'success': False,
            'error': '"quantity" должно быть числом'
        }, status=400)

    try:
        calculator = ComponentCalculator()
        result = calculator.calculate_for_product(article, quantity)
        return JsonResponse(result)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Ошибка расчёта: {str(e)}'
        }, status=500)


@require_http_methods(["POST"])
def export_components_excel(request):
    """
    Экспорт результатов расчёта в Excel.

    Request:
        POST application/json (result from calculate endpoints)
        OR
        POST multipart/form-data with file (recalculates and exports)

    Response:
        Excel file (.xlsx)
    """
    try:
        # Если передан файл - пересчитываем
        if 'file' in request.FILES:
            file = request.FILES['file']
            if not file.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({
                    'success': False,
                    'error': 'Неверный формат файла'
                }, status=400)
            calculator = ComponentCalculator()
            result = calculator.calculate_from_excel(file.read())
        else:
            # Иначе берём JSON из тела запроса
            try:
                result = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({
                    'success': False,
                    'error': 'Неверный формат данных'
                }, status=400)

        if not result.get('success', True):
            return JsonResponse(result, status=400)

        # Создаём Excel файл
        wb = Workbook()

        # Стили
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Лист 1: Сводка компонентов
        ws_summary = wb.active
        ws_summary.title = "Сводка компонентов"

        headers = ['№', 'Наименование материала', 'Ед.изм.', 'Количество']
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        components = result.get('components_summary', [])
        for row, comp in enumerate(components, 2):
            ws_summary.cell(row=row, column=1, value=row-1).border = thin_border
            ws_summary.cell(row=row, column=2, value=comp['material_name']).border = thin_border
            ws_summary.cell(row=row, column=3, value=comp['uom']).border = thin_border
            cell = ws_summary.cell(row=row, column=4, value=comp['total_quantity'])
            cell.border = thin_border
            cell.number_format = '#,##0.00'

        # Ширина колонок
        ws_summary.column_dimensions['A'].width = 6
        ws_summary.column_dimensions['B'].width = 60
        ws_summary.column_dimensions['C'].width = 10
        ws_summary.column_dimensions['D'].width = 15

        # Лист 2: Детали по товарам
        ws_products = wb.create_sheet("Детали по товарам")

        headers = ['Артикул', 'Наименование товара', 'Кол-во', 'Техкарта', 'Статус']
        for col, header in enumerate(headers, 1):
            cell = ws_products.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        products = result.get('products', [])
        for row, prod in enumerate(products, 2):
            status = 'Рассчитан' if prod.get('has_processing_plan') else (
                'Нет техкарты' if prod.get('found') else 'Не найден'
            )
            ws_products.cell(row=row, column=1, value=prod['article']).border = thin_border
            ws_products.cell(row=row, column=2, value=prod['name']).border = thin_border
            ws_products.cell(row=row, column=3, value=prod['quantity']).border = thin_border
            ws_products.cell(row=row, column=4, value=prod.get('processing_plan_name', '-')).border = thin_border
            ws_products.cell(row=row, column=5, value=status).border = thin_border

        ws_products.column_dimensions['A'].width = 15
        ws_products.column_dimensions['B'].width = 60
        ws_products.column_dimensions['C'].width = 10
        ws_products.column_dimensions['D'].width = 60
        ws_products.column_dimensions['E'].width = 15

        # Лист 3: Статистика
        ws_stats = wb.create_sheet("Статистика")
        stats = [
            ('Товаров найдено', result.get('products_found', 0)),
            ('Товаров не найдено', result.get('products_not_found', 0)),
            ('Товаров без техкарт', result.get('products_without_processing_plan', 0)),
            ('Всего компонентов', len(components)),
        ]
        for row, (label, value) in enumerate(stats, 1):
            ws_stats.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws_stats.cell(row=row, column=2, value=value)

        ws_stats.column_dimensions['A'].width = 25
        ws_stats.column_dimensions['B'].width = 15

        # Сохраняем в буфер
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=production_components.xlsx'
        return response

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Ошибка экспорта: {str(e)}'
        }, status=500)


@require_http_methods(["GET"])
def search_products_with_recipes(request):
    """
    Поиск товаров для расчёта производства.

    Query params:
        - search: строка поиска (по названию или артикулу)
        - limit: максимальное количество результатов (по умолчанию 20)

    Response:
        {
            "success": bool,
            "products": [
                {
                    "id": int,
                    "article": str,
                    "name": str,
                    "has_processing_plan": bool
                }
            ]
        }
    """
    from core.models import Product, ProcessingPlan, ProcessingPlanProduct
    from django.db.models import Q, Exists, OuterRef

    search = request.GET.get('search', '').strip()
    limit = min(int(request.GET.get('limit', 20)), 100)

    if not search:
        return JsonResponse({
            'success': True,
            'products': []
        })

    try:
        # Поиск товаров с аннотацией наличия техкарты (только с артикулами)
        products = Product.objects.filter(
            group='Товары'
        ).exclude(
            article__isnull=True
        ).exclude(
            article=''
        ).filter(
            Q(name__icontains=search) | Q(article__icontains=search)
        ).annotate(
            has_processing_plan=Exists(
                ProcessingPlanProduct.objects.filter(product=OuterRef('pk'))
            )
        ).order_by('-has_processing_plan', 'name')[:limit]

        result = [{
            'id': p.id,
            'article': p.article or '',
            'name': p.name,
            'has_processing_plan': p.has_processing_plan
        } for p in products]

        return JsonResponse({
            'success': True,
            'products': result
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
