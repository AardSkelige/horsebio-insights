# ozon/services/stock_analyzer.py
import pandas as pd
import numpy as np
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from io import BytesIO

import logging

logger = logging.getLogger(__name__)

def process_availability_data(input_file):
    """Обрабатывает файл с данными о доступности товаров и возвращает содержимое Excel-файла как bytes"""
    logger.info(f"Обработка файла: {input_file}")
    
    # Определим сначала метаданные из первых строк
    meta_df = pd.read_excel(input_file, header=None, nrows=4)
    
    # Читаем файл, используя строку 4 в качестве заголовка (индекс 4, 5-я строка в Excel)
    header_row = 4
    df = pd.read_excel(input_file, header=header_row)
    
    logger.info(f"Успешно прочитан файл. Количество строк: {df.shape[0]}, колонок: {df.shape[1]}")
    
    # Получение уникальных значений кластера для обработки данных по регионам
    clusters = df['кластер'].unique()
    logger.info(f"Уникальные кластеры: {clusters}")
    
    # Сначала получим уникальные наименования и артикулы
    unique_items = df[['имя (необязательно)', 'артикул']].drop_duplicates()
    
    # Создадим базовую структуру для нового DataFrame
    result_columns = [
        'Наименование', 'Артикул', 'Остаток суммарно', 'В пути суммарно'
    ]
    
    # Добавим колонки для каждого региона
    regions = sorted(df['кластер'].unique())
    for region in regions:
        result_columns.extend([
            f'{region} остаток, шт', 
            f'{region} в пути, шт', 
            f'{region} среднесут. продажи, шт'
        ])
    
    result_df = pd.DataFrame(columns=result_columns)
    
    # Заполним DataFrame данными
    item_idx = 0
    for _, row in unique_items.iterrows():
        item_name = row['имя (необязательно)']
        item_code = row['артикул']
        
        new_row = {'Наименование': item_name, 'Артикул': item_code}
        
        # Суммарные значения
        item_data = df[df['артикул'] == item_code]
        total_stock = item_data['остаток, шт'].sum()
        total_transit = item_data['товары в пути, шт'].sum()
        
        new_row['Остаток суммарно'] = total_stock
        new_row['В пути суммарно'] = total_transit
        
        # Данные по регионам
        for _, item_row in item_data.iterrows():
            region = item_row['кластер']
            
            new_row[f'{region} остаток, шт'] = item_row['остаток, шт']
            new_row[f'{region} в пути, шт'] = item_row['товары в пути, шт']
            new_row[f'{region} среднесут. продажи, шт'] = item_row['среднесут. продажи, шт']
        
        result_df.loc[item_idx] = new_row
        item_idx += 1
    
    # Заполним пустые ячейки нулями
    result_df.fillna(0, inplace=True)
    
    # Создаем буфер в памяти для сохранения файла Excel
    output = BytesIO()
    
    # Используем ExcelWriter для создания форматированного файла
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Записываем метаданные
        meta_df.to_excel(writer, sheet_name='Sheet1', header=False, index=False)
        
        # Добавляем цветовую легенду
        legend_df = pd.DataFrame([
            ['Расцветка:'],
            ['0-5, шт', '6-30, шт', '30-1000,шт']
        ])
        legend_df.to_excel(writer, sheet_name='Sheet1', header=False, index=False, startrow=1, startcol=6)
        
        # Записываем данные
        result_df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=4)
    
    # Применяем дополнительное форматирование
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active
    
    # Словарь с комментариями к колонкам
    column_comments = {
        'Наименование': 'Название товара из исходного файла (поле "имя (необязательно)")',
        'Артикул': 'Артикул товара из исходного файла',
        'Остаток суммарно': 'Сумма остатков по всем регионам (из поля "остаток, шт")',
        'В пути суммарно': 'Сумма товаров в пути по всем регионам (из поля "товары в пути, шт")',
    }
    
    # Добавляем комментарии для региональных колонок
    regions = []
    for col in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=5, column=col)
        if header_cell.value and isinstance(header_cell.value, str):
            header_text = header_cell.value
            # Проверяем, содержит ли заголовок название региона и тип данных
            match = re.match(r'(.*?) (остаток, шт|в пути, шт|среднесут\. продажи, шт)', header_text)
            if match:
                region = match.group(1)
                data_type = match.group(2)
                if region not in regions:
                    regions.append(region)
                
                if data_type == 'остаток, шт':
                    column_comments[header_text] = f'Текущий остаток в регионе {region} (из поля "остаток, шт")'
                elif data_type == 'в пути, шт':
                    column_comments[header_text] = f'Количество товара в пути в регион {region} (из поля "товары в пути, шт")'
                elif data_type == 'среднесут. продажи, шт':
                    column_comments[header_text] = f'Среднесуточные продажи в регионе {region} (из поля "среднесут. продажи, шт")'
    
    # Форматирование заголовков (строка 5 в Excel, индекс 4)
    header_row_idx = 5
    header_font = Font(bold=True, color="000000")  # Черный текст для заголовков
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Стиль для границ
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Цвета для легенды - точные цвета
    legend_colors = {
        '0-5, шт': 'F4CCCC',     # Красный
        '6-30, шт': 'FFE599',    # Желтый
        '30-1000,шт': 'B6D7A8'   # Зеленый
    }
    
    # Комментарии для легенды
    legend_comments = {
        '0-5, шт': 'Критический уровень запасов - требуется срочное пополнение',
        '6-30, шт': 'Средний уровень запасов - скоро потребуется пополнение',
        '30-1000,шт': 'Оптимальный уровень запасов'
    }
    
    # Форматирование легенды и добавление комментариев
    for cell_value, color in legend_colors.items():
        for row in range(1, 4):
            for col in range(1, 15):
                cell = ws.cell(row=row, column=col)
                if cell.value == cell_value:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    if cell_value in legend_comments:
                        cell.comment = Comment(legend_comments[cell_value], "Пояснение")
                    break
    
# ozon/services/stock_analyzer.py (продолжение)

    # Цвета для чередующихся колонок (копируем с образца)
    column_colors = {
        'Наименование': 'B4C6E7',          # Светло-голубой
        'Артикул': 'B4C6E7',               # Светло-голубой
        'Остаток суммарно': 'FFEB9C',      # Желтый
        'В пути суммарно': 'D9D9D9',       # Серый
    }
    
    # Цвета для регионов - копируем с образца
    region_colors = {
        'Москва-Восток': 'F4CCCC',   # Красный
        'Москва-Запад': 'B4C6E7',    # Голубой
        'Санкт-Петербург': 'D5A6BD', # Розовый
        'Дон': 'FFD966',             # Желтый
        'Юг': 'F9CB9C',              # Оранжевый
        'Поволжье': '9FC5E8',        # Синий
        'Урал': 'B4A7D6',            # Фиолетовый
        'Сибирь': 'D5D5D5',          # Серый
        'Калининград': 'EA9999',     # Красный
        'Дальний Восток': 'A2C4C9',  # Бирюзовый
        'Беларусь': 'F4CCCC',        # Красный
        'Казахстан': 'D9D2E9'        # Светло-пурпурный
    }
    
    # Находим последнюю колонку
    max_col = ws.max_column
    
    # Форматирование заголовков и добавление комментариев
    for col in range(1, max_col + 1):
        cell = ws.cell(row=header_row_idx, column=col)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # Добавление комментариев к заголовкам
        if cell.value in column_comments:
            cell.comment = Comment(column_comments[cell.value], "Описание")
        
        # Установка цвета заголовка на основе типа колонки
        header_text = str(cell.value) if cell.value else ""
        
        # Проверка по имени колонки
        if header_text in column_colors:
            cell.fill = PatternFill(start_color=column_colors[header_text], end_color=column_colors[header_text], fill_type="solid")
        else:
            # Проверка по названию региона
            region_match = None
            for region in region_colors:
                if region in header_text:
                    region_match = region
                    break
            
            if region_match:
                # Чередование цветов для разных типов данных региона
                if "остаток, шт" in header_text:
                    cell.fill = PatternFill(start_color=region_colors[region_match], end_color=region_colors[region_match], fill_type="solid")
                elif "в пути, шт" in header_text:
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Серый для всех "в пути"
                elif "среднесут. продажи" in header_text:
                    cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Светло-голубой для продаж
            else:
                # Если не нашли подходящий цвет, используем стандартный
                cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    
    # Форматирование данных
    for row in range(header_row_idx + 1, header_row_idx + len(result_df) + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            
            header_cell = ws.cell(row=header_row_idx, column=col)
            header_text = str(header_cell.value) if header_cell.value else ""
            
            # Проверяем, является ли ячейка показателем остатка
            is_stock_cell = "остаток, шт" in header_text
            
            # Применяем цвет фона в зависимости от типа колонки
            if header_text in column_colors:
                # Для базовых колонок
                if header_text == "Остаток суммарно" and cell.value != 0:
                    # Для Остаток суммарно используем условное форматирование
                    value = cell.value if cell.value else 0
                    if 0 <= value < 6:
                        cell.fill = PatternFill(start_color=legend_colors['0-5, шт'], end_color=legend_colors['0-5, шт'], fill_type="solid")
                    elif 6 <= value < 31:
                        cell.fill = PatternFill(start_color=legend_colors['6-30, шт'], end_color=legend_colors['6-30, шт'], fill_type="solid")
                    elif value >= 31:
                        cell.fill = PatternFill(start_color=legend_colors['30-1000,шт'], end_color=legend_colors['30-1000,шт'], fill_type="solid")
                else:
                    # Для остальных базовых колонок
                    cell.fill = PatternFill(start_color=column_colors[header_text], end_color=column_colors[header_text], fill_type="solid")
            elif is_stock_cell and cell.value != 0:
                # Для колонок с остатками применяем условное форматирование
                value = cell.value if cell.value else 0
                if 0 <= value < 6:
                    cell.fill = PatternFill(start_color=legend_colors['0-5, шт'], end_color=legend_colors['0-5, шт'], fill_type="solid")
                elif 6 <= value < 31:
                    cell.fill = PatternFill(start_color=legend_colors['6-30, шт'], end_color=legend_colors['6-30, шт'], fill_type="solid")
                elif value >= 31:
                    cell.fill = PatternFill(start_color=legend_colors['30-1000,шт'], end_color=legend_colors['30-1000,шт'], fill_type="solid")
            elif "в пути, шт" in header_text and cell.value != 0:
                # Серый цвет для ячеек "в пути"
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            elif "среднесут. продажи" in header_text and cell.value != 0:
                # Светло-голубой цвет для ячеек продаж
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Автоподбор ширины колонок
    for col in range(1, max_col + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15
        
        # Для колонок "Наименование" и "Артикул" устанавливаем большую ширину
        cell = ws.cell(row=header_row_idx, column=col)
        if cell.value == "Наименование":
            ws.column_dimensions[column_letter].width = 40
        elif cell.value == "Артикул":
            ws.column_dimensions[column_letter].width = 20
    
    # Форматирование метаданных (заголовка отчета)
    for row in range(1, header_row_idx):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and "Дата формирования" in cell.value:
                cell.font = Font(bold=True, size=12)
                cell.comment = Comment("Дата создания отчета", "Информация")
            elif cell.value and isinstance(cell.value, str) and "Отчет за период" in cell.value:
                cell.font = Font(bold=True, size=12)
                cell.comment = Comment("Период времени, за который собраны данные", "Информация")
            elif cell.value and isinstance(cell.value, str) and "Рекомендуемая поставка" in cell.value:
                cell.font = Font(bold=True, size=12)
                cell.comment = Comment("Значение получено из исходного файла. Представляет собой рекомендуемый период планирования поставок.", "Информация")
            elif cell.value and cell.value == "Расцветка:":
                cell.font = Font(bold=True)
                cell.comment = Comment("Цветовая кодировка для уровней запасов", "Информация")
    
    # Закрепление заголовков при прокрутке
    freeze_cell = f'A{header_row_idx + 1}'
    ws.freeze_panes = freeze_cell
    
    # Сохраняем отформатированный файл в новый буфер
    formatted_output = BytesIO()
    wb.save(formatted_output)
    formatted_output.seek(0)
    
    return formatted_output.getvalue()
    