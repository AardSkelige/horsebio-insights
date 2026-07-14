# ozon/services/report_generator.py

import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Optional
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ReportGenerator:
    """
    ОБНОВЛЕННЫЙ генератор отчета DRR с точным воспроизведением структуры, 
    формул и форматирования как в эталонном файле DRR.xlsx
    
    Генерирует отчет на основе двух исходных файлов:
    - analytics_report_*.xlsx (данные по товарам из Ozon)  
    - sku_statistics_*.xlsx (данные по рекламе из Ozon)
    
    ИСПОЛЬЗОВАНИЕ:
    generator = ReportGenerator()
    excel_data = generator.generate_report(analytics_file, sku_file)
    """
    
    def __init__(self):
        """Инициализация с точными настройками эталонного файла"""
        
        # Точные названия колонок как в DRR.xlsx (строка 3)
        self.column_names = [
            'Наименование', 'Артикул', 'Трафареты расход, ₽', 'Вывод в топ расход, ₽', 
            'Оплата за заказ расход, ₽', 'Спецразмещение расход, ₽', 'Общий расход на все виды рекламы, ₽',
            'Заказанно на сумму, ₽', 'Заказано штук', 'Выкуп, %', 'Средняя цена одного выкупа, ₽', 'Дрр, %',
            'Трафареты кол-во показов, шт', 'Трафареты кол-во кликов, шт', 'Трафареты CTR, %',
            'Трафареты средняя цена клика, ₽', 'Трафареты конверсия в корзину, %', 'Трафареты кол-во корзин, шт',
            'Трафареты кол-во заказов, шт', 'Трафареты Конверсия из корзины в заказ, %',
            'Трафареты кол-во корзин для одного заказа, шт', 'Трафареты цена одной корзины, ₽',
            'Трафареты стоимость заказа, ₽', 'Трафареты заказано на сумму, ₽',
            'Вывод в топ кол-во показов, шт', 'Вывод в топ кол-во кликов, шт', 'Вывод в топ CTR, %',
            'Вывод в топ средняя цена клика, ₽', 'Вывод в топ конверсия в корзину, %', 'Вывод в топ кол-во корзин, шт',
            'Вывод в топ кол-во заказов, шт', 'Вывод в топ Конверсия из корзины в заказ, %',
            'Вывод в топ кол-во корзин для одного заказа, шт', 'Вывод в топ стоимость одной корзины, ₽',
            'Вывод в топ стоимость заказа, ₽', 'Вывод в топ заказано на сумму, ₽',
            'Оплата за заказ кол-во показов, шт', 'Оплата за заказ кол-во кликов, шт', 'Оплата за заказ CTR, %',
            'Оплата за заказ конверсия в корзину, %', 'Оплата за заказ кол-во корзин, шт', 'Оплата за заказ кол-во заказов, шт',
            'Оплата за заказ Конверсия из корзины в заказ %', 'Оплата за заказ кол-во корзин для одного заказа, шт',
            'Оплата за заказ стоимость одной корзины, ₽', 'Оплата за заказ средний расход на заказ, ₽',
            'Оплата за заказ заказано на сумму, ₽', 'Спецразмещение кол-во показов, шт', 'Спецразмещение кол-во кликов, шт',
            'Спецразмещение CTR, %', 'Спецразмещение средняя цена клика, ₽', 'Спецразмещение конверсия в корзину, %',
            'Спецразмещение кол-во корзин, шт', 'Спецразмещение кол-во заказов, шт', 'Спецразмещение Конверсия из корзины в заказ, %',
            'Спецразмещение кол-во корзин для одного заказа, шт', 'Спецразмещение стоимость одной корзины, ₽',
            'Спецразмещение стоимость заказа, ₽', 'Спецразмещение заказано на сумму, ₽',
            'Кол-во показов в рекламе суммарно, шт', 'Кол-во кликов в рекламе суммарно, шт',
            'Уникальные посетители, всего', 'Уникальные посетители с просмотром карточки товара',
            'Конверсия в корзину из карточки товара %', 'Кол-во добавлений в корзину (в штуках)',
            'Заказано штук', 'Конверсия из корзины в заказ %', 'Кол-во корзин для одного заказа, шт',
            'Стоимость добавления в корзину, ₽', 'Кол-во добавлений в корзину по рекламе, шт', 'Органические корзины, шт',
            'Средняя цена товара, ₽', 'Кликов для одного выкупа, шт', 'Желаемый Дрр %',
            'Сколько готовы тратить на рекламу в рублях, ₽', 'Средняя Конверсия из корзины в заказ, %',
            'Средняя цена корзины, для попадания в ДРР, ₽', 'Если стоимость корзины будет такой, ₽',
            'Тратить на рекламу для одного заказа буду, ₽', 'Примерный дрр, %'
        ]
        
    def _is_analytics_file(self, df):
        """Определяет, является ли DataFrame файлом analytics по наличию характерных столбцов"""
        if df is None or df.empty:
            return False
        
        analytics_indicators = ['Товары', 'Ozon ID', 'Артикул', 'Заказано товаров', 'Уникальные посетители']
        sku_indicators = ['SKU', 'Тип продвижения', 'Расход', 'Продажи']
        
        # Проверяем наличие характерных столбцов для analytics
        analytics_score = sum(1 for col in df.columns if any(indicator in str(col) for indicator in analytics_indicators))
        sku_score = sum(1 for col in df.columns if any(indicator in str(col) for indicator in sku_indicators))
        
        logger.info(f"🔍 Анализ файла - Analytics маркеры: {analytics_score}, SKU маркеры: {sku_score}")
        
        return analytics_score > sku_score

    def clean_number(self, x):
        """Очистка и преобразование числовых значений"""
        if pd.isna(x) or x == '' or x is None:
            return 0
        if isinstance(x, str):
            cleaned = str(x).replace(' ', '').replace(',', '.').replace('%', '').replace('₽', '').replace('—', '0')
            try:
                return float(cleaned)
            except ValueError:
                return 0
        return float(x)

    def load_source_data(self, file1, file2):
        """Загрузка исходных данных из файлов или файловых объектов"""
        logger.info("📥 ЗАГРУЖАЕМ ИСХОДНЫЕ ДАННЫЕ")
        
        try:
            # Определяем тип входных данных (путь к файлу или файловый объект)
            if hasattr(file1, 'read'):
                # Это файловые объекты Django
                logger.info("📁 Загружаем из файловых объектов Django")
                
                # Читаем оба файла для анализа
                df1_default = pd.read_excel(file1)
                df1_header1 = pd.read_excel(file1, header=1) 
                file1.seek(0)  # Возвращаемся к началу
                
                df2_default = pd.read_excel(file2)
                df2_header1 = pd.read_excel(file2, header=1)
                file2.seek(0)  # Возвращаемся к началу
                
                # Проверяем какой из способов чтения дает нужную структуру
                logger.info("🔍 Анализируем все варианты чтения файлов...")
                
                # Определяем правильное сочетание файлов и способов чтения
                if self._is_analytics_file(df1_default):
                    logger.info("🔍 file1 (default) = Analytics, file2 (header=1) = SKU")
                    df_analytics = df1_default
                    df_sku = df2_header1
                elif self._is_analytics_file(df1_header1):
                    logger.info("🔍 file1 (header=1) = Analytics, file2 (default) = SKU") 
                    df_analytics = df1_header1
                    df_sku = df2_default
                elif self._is_analytics_file(df2_default):
                    logger.info("🔍 file2 (default) = Analytics, file1 (header=1) = SKU")
                    df_analytics = df2_default
                    df_sku = df1_header1
                elif self._is_analytics_file(df2_header1):
                    logger.info("🔍 file2 (header=1) = Analytics, file1 (default) = SKU")
                    df_analytics = df2_header1
                    df_sku = df1_default
                else:
                    logger.warning("⚠️ Не удалось определить файлы, пробуем ads_file как SKU, products_file как Analytics")
                    # По названию: ads_file должен быть SKU, products_file - Analytics
                    df_sku = df1_header1  # ads_file как SKU с header=1
                    df_analytics = df2_default  # products_file как Analytics
            else:
                # Это пути к файлам
                logger.info("📁 Загружаем из файловых путей")
                df_analytics = pd.read_excel(file1)
                df_sku = pd.read_excel(file2, header=1)
            
            # Если analytics файл все еще не распознан, пробуем другие варианты чтения
            if not self._is_analytics_file(df_analytics):
                logger.info("📋 Пробуем разные способы чтения analytics файла...")
                if hasattr(file1, 'read'):
                    file1.seek(0)
                    try:
                        df_analytics_alt = pd.read_excel(file1, header=1)
                        if self._is_analytics_file(df_analytics_alt):
                            df_analytics = df_analytics_alt
                    except Exception:
                        pass
            
            logger.info(f"📈 Analytics загружен: {len(df_analytics)} товаров")
            logger.info(f"📋 Analytics columns: {list(df_analytics.columns)}")
            
            # Показываем первые строки analytics для отладки
            logger.info("📋 Analytics первые 3 строки:")
            for i in range(min(3, len(df_analytics))):
                logger.info(f"   Строка {i}: {df_analytics.iloc[i].tolist()[:5]}")  # Первые 5 колонок
            
            logger.info(f"📊 SKU загружен: {len(df_sku)} записей")
            logger.info(f"📋 SKU columns: {list(df_sku.columns)}")
            
            # Показываем первые строки SKU для отладки
            logger.info("📋 SKU первые 3 строки:")
            for i in range(min(3, len(df_sku))):
                logger.info(f"   Строка {i}: {df_sku.iloc[i].tolist()[:5]}")  # Первые 5 колонок
            
            return df_analytics, df_sku
            
        except Exception as e:
            logger.error(f"❌ Ошибка загрузки: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None, None

    def create_product_data_mapping(self, df_analytics, df_sku):
        """Создание маппинга данных по артикулам"""
        logger.info("📋 СОЗДАЕМ МАППИНГ ДАННЫХ ПО АРТИКУЛАМ")

        # Словарь для хранения всех данных по артикулам
        product_data = {}

        # Определяем название столбца с Ozon ID
        ozon_id_column = None
        possible_ozon_columns = ['Ozon ID', 'OZON ID', 'ozon_id', 'ID', 'id', 'SKU']
        for col in possible_ozon_columns:
            if col in df_analytics.columns:
                ozon_id_column = col
                break

        if not ozon_id_column:
            logger.error(f"❌ Не найден столбец с Ozon ID в analytics файле. Доступные столбцы: {list(df_analytics.columns)}")
            return {}

        logger.info(f"📌 Используем столбец '{ozon_id_column}' как Ozon ID")

        # Обрабатываем analytics_report
        for idx, row in df_analytics.iterrows():
            ozon_id = str(row[ozon_id_column])
            article = str(row['Артикул'])

            # Безопасное извлечение данных с проверкой наличия столбцов
            def safe_get_column(column_names, default=0):
                for col_name in column_names if isinstance(column_names, list) else [column_names]:
                    if col_name in df_analytics.columns:
                        return self.clean_number(row[col_name])
                logger.warning(f"❌ Столбец не найден: {column_names}")
                return default

            product_data[article] = {
                'ozon_id': ozon_id,
                'name': row.get('Товары', row.get('Наименование', row.get('Название', ''))),
                'ordered_amount': safe_get_column(['Заказано на сумму', 'Сумма заказов', 'Revenue']),
                'ordered_qty': safe_get_column(['Заказано товаров', 'Количество заказов', 'Orders']),
                'unique_visitors': safe_get_column(['Уникальные посетители, всего', 'Посетители']),
                'unique_visitors_product': safe_get_column(['Уникальные посетители с просмотром карточки товара', 'Посетители карточки']),
                'conversion_to_cart': safe_get_column(['Конверсия в корзину из карточки товара', 'Конверсия в корзину']),

                # Инициализируем рекламные данные нулями
                'ad_data': {
                    'Трафареты': {'spend': 0, 'shows': 0, 'clicks': 0, 'baskets': 0, 'orders': 0, 'sales': 0},
                    'Вывод в топ': {'spend': 0, 'shows': 0, 'clicks': 0, 'baskets': 0, 'orders': 0, 'sales': 0},
                    'Оплата за заказ': {'spend': 0, 'shows': 0, 'clicks': 0, 'baskets': 0, 'orders': 0, 'sales': 0},
                    'Спецразмещение': {'spend': 0, 'shows': 0, 'clicks': 0, 'baskets': 0, 'orders': 0, 'sales': 0}
                }
            }

        # Определяем формат SKU файла (старый или новый)
        is_new_format = 'Инструмент' in df_sku.columns
        logger.info(f"📊 Обнаружен {'новый' if is_new_format else 'старый'} формат SKU файла")

        # Обрабатываем sku_statistics
        for idx, row in df_sku.iterrows():
            sku = str(row['SKU'])

            # Получаем тип рекламы в зависимости от формата
            if is_new_format:
                ad_type = row['Инструмент']  # Новый формат
            else:
                ad_type = row['Тип продвижения']  # Старый формат

            # Находим соответствующий артикул по SKU = Ozon ID
            matching_article = None
            for article, data in product_data.items():
                if data['ozon_id'] == sku:
                    matching_article = article
                    break

            if matching_article:
                # Нормализуем тип рекламы
                normalized_ad_type = ad_type
                if ad_type == 'Продвижение в поиске':
                    normalized_ad_type = 'Оплата за заказ'

                if normalized_ad_type in product_data[matching_article]['ad_data']:
                    ad_data = product_data[matching_article]['ad_data'][normalized_ad_type]

                    # Извлекаем данные в зависимости от формата
                    if is_new_format:
                        # Новый формат - другие позиции колонок
                        ad_data['spend'] += self.clean_number(row['Расход, ₽'])
                        ad_data['shows'] += self.clean_number(row['Показы'])
                        ad_data['clicks'] += self.clean_number(row['Клики'])
                        ad_data['baskets'] += self.clean_number(row['В корзину'])
                        ad_data['orders'] += self.clean_number(row['Заказы, шт'])
                        ad_data['sales'] += self.clean_number(row['Продажи, ₽'])
                    else:
                        # Старый формат
                        ad_data['spend'] += self.clean_number(row['Расход, ₽, с НДС'])
                        ad_data['shows'] += self.clean_number(row['Показы'])
                        ad_data['clicks'] += self.clean_number(row['Клики'])
                        ad_data['baskets'] += self.clean_number(row['Корзины'])
                        ad_data['orders'] += self.clean_number(row['Заказы, шт'])
                        ad_data['sales'] += self.clean_number(row['Продажи, ₽'])

        logger.info(f"✅ Маппинг создан для {len(product_data)} товаров")
        return product_data

    def create_excel_with_exact_structure(self, product_data):
        """Создание Excel файла с точной структурой DRR"""
        logger.info("📊 СОЗДАЕМ EXCEL С ТОЧНОЙ СТРУКТУРОЙ DRR")
        
        # Создаем workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # СТРОКА 1: Пустая (как в эталоне)
        
        # СТРОКА 2: Итоговая строка с заголовком и формулами SUBTOTAL
        ws.cell(row=2, column=1).value = "Сумма в стобцах и общий ДРР:"
        
        # СТРОКА 3: Заголовки
        for col_idx, col_name in enumerate(self.column_names, 1):
            ws.cell(row=3, column=col_idx).value = col_name
        
        # СТРОКИ 4+: Данные товаров
        row_idx = 4
        for article, data in product_data.items():
            # Базовые данные
            ws.cell(row=row_idx, column=1).value = data['name']  # A - Наименование
            ws.cell(row=row_idx, column=2).value = article  # B - Артикул
            
            # Расходы на рекламу (прямые значения)
            ws.cell(row=row_idx, column=3).value = data['ad_data']['Трафареты']['spend']  # C - Трафареты расход
            ws.cell(row=row_idx, column=4).value = data['ad_data']['Вывод в топ']['spend']  # D - Вывод в топ расход
            ws.cell(row=row_idx, column=5).value = data['ad_data']['Оплата за заказ']['spend']  # E - Оплата за заказ расход
            ws.cell(row=row_idx, column=6).value = data['ad_data']['Спецразмещение']['spend']  # F - Спецразмещение расход
            
            # G - Общий расход (формула)
            ws.cell(row=row_idx, column=7).value = f'=SUM(C{row_idx}:F{row_idx})'
            
            # Основные показатели
            ws.cell(row=row_idx, column=8).value = data['ordered_amount']  # H - Заказанно на сумму
            ws.cell(row=row_idx, column=9).value = data['ordered_qty']  # I - Заказано штук
            ws.cell(row=row_idx, column=10).value = 95  # J - Выкуп (константа)
            
            # K - Средняя цена одного выкупа (формула)
            ws.cell(row=row_idx, column=11).value = f'=IFERROR(G{row_idx}/(I{row_idx}*0.01*J{row_idx}), 0)'
            
            # L - ДРР (формула)  
            ws.cell(row=row_idx, column=12).value = f'=IFERROR(IF(AND(G{row_idx} > 0,H{row_idx}=0), 100, G{row_idx}/(H{row_idx}*0.01*J{row_idx})*100), 0)'
            
            # Данные по трафаретам
            ws.cell(row=row_idx, column=13).value = data['ad_data']['Трафареты']['shows']  # M - Трафареты показы
            ws.cell(row=row_idx, column=14).value = data['ad_data']['Трафареты']['clicks']  # N - Трафареты клики
            ws.cell(row=row_idx, column=15).value = f'=IFERROR(N{row_idx}/M{row_idx}*100, 0)'  # O - CTR
            ws.cell(row=row_idx, column=16).value = f'=IFERROR(C{row_idx}/N{row_idx}, 0)'  # P - Средняя цена клика
            ws.cell(row=row_idx, column=17).value = f'=IFERROR(R{row_idx}/N{row_idx}*100, 0)'  # Q - Конверсия в корзину
            ws.cell(row=row_idx, column=18).value = data['ad_data']['Трафареты']['baskets']  # R - Корзины
            ws.cell(row=row_idx, column=19).value = data['ad_data']['Трафареты']['orders']  # S - Заказы
            ws.cell(row=row_idx, column=20).value = f'=IFERROR(S{row_idx}/R{row_idx}*100, 0)'  # T - Конверсия из корзины в заказ
            ws.cell(row=row_idx, column=21).value = f'=IFERROR(R{row_idx}/S{row_idx}, 0)'  # U - Корзин на заказ
            ws.cell(row=row_idx, column=22).value = f'=IFERROR(C{row_idx}/R{row_idx}, 0)'  # V - Цена корзины
            ws.cell(row=row_idx, column=23).value = f'=IFERROR(C{row_idx}/S{row_idx}, 0)'  # W - Стоимость заказа
            
            # X - Трафареты заказано на сумму
            ws.cell(row=row_idx, column=24).value = data['ad_data']['Трафареты']['sales']
            
            # Данные по "Вывод в топ"
            ws.cell(row=row_idx, column=25).value = data['ad_data']['Вывод в топ']['shows']  # Y - Показы
            ws.cell(row=row_idx, column=26).value = data['ad_data']['Вывод в топ']['clicks']  # Z - Клики
            ws.cell(row=row_idx, column=27).value = f'=IFERROR(Z{row_idx}/Y{row_idx}*100, 0)'  # AA - CTR
            ws.cell(row=row_idx, column=28).value = f'=IFERROR(D{row_idx}/Z{row_idx}, 0)'  # AB - Средняя цена клика
            ws.cell(row=row_idx, column=29).value = f'=IFERROR(AD{row_idx}/Z{row_idx}*100, 0)'  # AC - Конверсия в корзину
            ws.cell(row=row_idx, column=30).value = data['ad_data']['Вывод в топ']['baskets']  # AD - Корзины
            ws.cell(row=row_idx, column=31).value = data['ad_data']['Вывод в топ']['orders']  # AE - Заказы
            ws.cell(row=row_idx, column=32).value = f'=IFERROR(AE{row_idx}/AD{row_idx}*100, 0)'  # AF - Конверсия из корзины
            ws.cell(row=row_idx, column=33).value = f'=IFERROR(AD{row_idx}/AE{row_idx}, 0)'  # AG - Корзин на заказ
            ws.cell(row=row_idx, column=34).value = f'=IFERROR(D{row_idx}/AD{row_idx}, 0)'  # AH - Стоимость корзины
            ws.cell(row=row_idx, column=35).value = f'=IFERROR(D{row_idx}/AE{row_idx}, 0)'  # AI - Стоимость заказа
            ws.cell(row=row_idx, column=36).value = data['ad_data']['Вывод в топ']['sales']  # AJ - Заказано на сумму
            
            # Данные по "Оплата за заказ"
            ws.cell(row=row_idx, column=37).value = data['ad_data']['Оплата за заказ']['shows']  # AK - Показы
            ws.cell(row=row_idx, column=38).value = data['ad_data']['Оплата за заказ']['clicks']  # AL - Клики
            ws.cell(row=row_idx, column=39).value = f'=IFERROR(AL{row_idx}/AK{row_idx}*100, 0)'  # AM - CTR
            ws.cell(row=row_idx, column=40).value = f'=IFERROR(AO{row_idx}/AL{row_idx}*100, 0)'  # AN - Конверсия в корзину
            ws.cell(row=row_idx, column=41).value = data['ad_data']['Оплата за заказ']['baskets']  # AO - Корзины
            ws.cell(row=row_idx, column=42).value = data['ad_data']['Оплата за заказ']['orders']  # AP - Заказы
            ws.cell(row=row_idx, column=43).value = f'=IFERROR(AP{row_idx}/AO{row_idx}*100, 0)'  # AQ - Конверсия из корзины
            ws.cell(row=row_idx, column=44).value = f'=IFERROR(AO{row_idx}/AP{row_idx}, 0)'  # AR - Корзин на заказ
            ws.cell(row=row_idx, column=45).value = f'=IFERROR(E{row_idx}/AO{row_idx}, 0)'  # AS - Стоимость корзины
            ws.cell(row=row_idx, column=46).value = f'=IFERROR(E{row_idx}/AP{row_idx}, 0)'  # AT - Средний расход на заказ
            ws.cell(row=row_idx, column=47).value = data['ad_data']['Оплата за заказ']['sales']  # AU - Заказано на сумму
            
            # Данные по "Спецразмещение"
            ws.cell(row=row_idx, column=48).value = data['ad_data']['Спецразмещение']['shows']  # AV - Показы
            ws.cell(row=row_idx, column=49).value = data['ad_data']['Спецразмещение']['clicks']  # AW - Клики
            ws.cell(row=row_idx, column=50).value = f'=IFERROR(AW{row_idx}/AV{row_idx}*100, 0)'  # AX - CTR
            ws.cell(row=row_idx, column=51).value = f'=IFERROR(F{row_idx}/AW{row_idx}, 0)'  # AY - Средняя цена клика
            ws.cell(row=row_idx, column=52).value = f'=IFERROR(BA{row_idx}/AW{row_idx}*100, 0)'  # AZ - Конверсия в корзину
            ws.cell(row=row_idx, column=53).value = data['ad_data']['Спецразмещение']['baskets']  # BA - Корзины
            ws.cell(row=row_idx, column=54).value = data['ad_data']['Спецразмещение']['orders']  # BB - Заказы
            ws.cell(row=row_idx, column=55).value = f'=IFERROR(BB{row_idx}/BA{row_idx}*100, 0)'  # BC - Конверсия из корзины
            ws.cell(row=row_idx, column=56).value = f'=IFERROR(BA{row_idx}/BB{row_idx}, 0)'  # BD - Корзин на заказ
            ws.cell(row=row_idx, column=57).value = f'=IFERROR(F{row_idx}/BA{row_idx}, 0)'  # BE - Стоимость корзины
            ws.cell(row=row_idx, column=58).value = f'=IFERROR(F{row_idx}/BB{row_idx}, 0)'  # BF - Стоимость заказа
            ws.cell(row=row_idx, column=59).value = data['ad_data']['Спецразмещение']['sales']  # BG - Заказано на сумму
            
            # Суммарные показы и клики
            ws.cell(row=row_idx, column=60).value = f'=IFERROR(M{row_idx}+Y{row_idx}+AK{row_idx}+AV{row_idx}, 0)'  # BH - Общие показы
            ws.cell(row=row_idx, column=61).value = f'=IFERROR(N{row_idx}+Z{row_idx}+AL{row_idx}+AW{row_idx}, 0)'  # BI - Общие клики
            
            # Данные из analytics
            ws.cell(row=row_idx, column=62).value = data['unique_visitors']  # BJ - Уникальные посетители
            ws.cell(row=row_idx, column=63).value = data['unique_visitors_product']  # BK - Уникальные с просмотром
            ws.cell(row=row_idx, column=64).value = data['conversion_to_cart']  # BL - Конверсия в корзину из карточки
            
            # Формульные столбцы
            ws.cell(row=row_idx, column=65).value = f'=IFERROR(BK{row_idx}*0.01*BL{row_idx}, 0)'  # BM - Добавлений в корзину
            ws.cell(row=row_idx, column=66).value = data['ordered_qty']  # BN - Заказано штук (дублирует I)
            ws.cell(row=row_idx, column=67).value = f'=IFERROR(BN{row_idx}/BM{row_idx}*100, 0)'  # BO - Конверсия из корзины
            ws.cell(row=row_idx, column=68).value = f'=IFERROR(BM{row_idx}/BN{row_idx}, 0)'  # BP - Корзин на заказ
            ws.cell(row=row_idx, column=69).value = f'=IFERROR(G{row_idx}/BM{row_idx}, 0)'  # BQ - Стоимость добавления в корзину
            ws.cell(row=row_idx, column=70).value = f'=IFERROR(R{row_idx}+AD{row_idx}+AO{row_idx}+BA{row_idx}, 0)'  # BR - Корзины по рекламе
            ws.cell(row=row_idx, column=71).value = f'=IFERROR(BM{row_idx}-BR{row_idx}, 0)'  # BS - Органические корзины
            ws.cell(row=row_idx, column=72).value = f'=IFERROR(H{row_idx}/I{row_idx}, 0)'  # BT - Средняя цена товара
            ws.cell(row=row_idx, column=73).value = f'=IFERROR(BK{row_idx}/(BN{row_idx}*0.01*J{row_idx}), 0)'  # BU - Кликов на выкуп
            
            # Планирование
            ws.cell(row=row_idx, column=74).value = 15  # BV - Желаемый ДРР (константа)
            ws.cell(row=row_idx, column=75).value = f'=IFERROR(BT{row_idx}*0.01*BV{row_idx}, 0)'  # BW - Готовы тратить
            ws.cell(row=row_idx, column=76).value = 20  # BX - Средняя конверсия (константа)
            ws.cell(row=row_idx, column=77).value = f'=IFERROR(BW{row_idx}/(100/BX{row_idx}), 0)'  # BY - Средняя цена корзины
            ws.cell(row=row_idx, column=78).value = 55  # BZ - Стоимость корзины (константа)
            ws.cell(row=row_idx, column=79).value = f'=IFERROR(BZ{row_idx}*(100/BX{row_idx}), 0)'  # CA - Тратить на заказ
            ws.cell(row=row_idx, column=80).value = f'=IFERROR(CA{row_idx}/BT{row_idx}*100, 0)'  # CB - Примерный ДРР
            
            row_idx += 1
        
        # Создаем формулы итоговой строки
        self.create_total_formulas(ws, len(product_data))
        
        # Применяем форматирование к данным товаров (2 знака после запятой)
        self.apply_data_formatting(ws, len(product_data))
        
        # Применяем форматирование заголовков
        self.apply_formatting(ws)
        
        # Принудительно включаем автоматический пересчет формул
        try:
            wb.calculation.calcMode = 'auto'
            wb.calculation.fullCalcOnLoad = True
        except Exception:
            # Если не удается установить режим расчета, продолжаем без него
            pass
        
        return wb

    def create_total_formulas(self, ws, num_products):
        """Создание формул итоговой строки"""
        # Формулы SUBTOTAL для сумм
        subtotal_columns = [3, 4, 5, 6, 7, 8, 9, 13, 14, 18, 19, 25, 26, 30, 31, 37, 38, 41, 42, 48, 49, 53, 54, 62, 63, 65, 66]
        
        for col in subtotal_columns:
            col_letter = get_column_letter(col)
            ws.cell(row=2, column=col).value = f'=SUBTOTAL(9,{col_letter}4:{col_letter}200000)'
        
        # Специальные формулы для итогов
        ws.cell(row=2, column=10).value = 95  # J2 - Выкуп константа
        ws.cell(row=2, column=11).value = '=IFERROR(G2/(I2*0.01*J2),0)'  # K2 - Средняя цена выкупа
        ws.cell(row=2, column=12).value = '=IFERROR(G2/(H2*0.01*J2)*100,0)'  # L2 - ДРР
        
        # CTR и другие расчетные
        ws.cell(row=2, column=15).value = '=IFERROR(N2/M2*100,0)'  # O2 - CTR трафареты
        ws.cell(row=2, column=16).value = '=IFERROR(C2/N2,0)'  # P2 - Средняя цена клика
        ws.cell(row=2, column=17).value = '=IFERROR(R2/N2*100,0)'  # Q2 - Конверсия в корзину
        ws.cell(row=2, column=20).value = '=IFERROR(S2/R2*100,0)'  # T2 - Конверсия из корзины
        
        # Аналогично для остальных типов рекламы
        ws.cell(row=2, column=27).value = '=IFERROR(Z2/Y2*100,0)'  # AA2 - CTR вывод в топ
        ws.cell(row=2, column=39).value = '=IFERROR(AL2/AK2*100,0)'  # AM2 - CTR оплата за заказ
        ws.cell(row=2, column=50).value = '=IFERROR(AW2/AV2*100,0)'  # AX2 - CTR спецразмещение
        
        # Суммы показов и кликов
        ws.cell(row=2, column=60).value = '=SUM(M2,Y2,AK2,AV2)'  # BH2 - Общие показы
        ws.cell(row=2, column=61).value = '=SUM(Z2,AL2,AW2,N2)'  # BI2 - Общие клики
        
        # Конверсии и другие расчеты
        ws.cell(row=2, column=64).value = '=IFERROR(BM2/BK2*100,0)'  # BL2 - Конверсия в корзину
        ws.cell(row=2, column=67).value = '=IFERROR(BN2/BM2*100,0)'  # BO2 - Конверсия из корзины
        ws.cell(row=2, column=69).value = '=IFERROR(G2/BM2,0)'  # BQ2 - Стоимость добавления в корзину
        ws.cell(row=2, column=70).value = '=SUM(R2,AD2,AO2,BA2)'  # BR2 - Корзины по рекламе
        ws.cell(row=2, column=72).value = '=IFERROR(H2/I2,0)'  # BT2 - Средняя цена товара
        ws.cell(row=2, column=73).value = '=IFERROR(BK2/(BN2*0.01*J2),0)'  # BU2 - Кликов на выкуп

    def apply_formatting(self, ws):
        """Применение форматирования как в эталоне"""
        
        # Стили для заголовков (строка 3)
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF2F75B5', end_color='FF2F75B5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Стили для итоговой строки (строка 2)
        total_font = Font(name='Calibri', size=11, bold=True)
        total_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Применяем к заголовкам
        for col in range(1, 81):  # 80 колонок
            # Заголовки
            header_cell = ws.cell(row=3, column=col)
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = header_alignment
            
            # Итоговая строка
            total_cell = ws.cell(row=2, column=col)
            total_cell.font = total_font
            total_cell.alignment = total_alignment
            
            # Форматы чисел - все до 2 знаков после запятой
            if 'расход' in self.column_names[col-1].lower() or '₽' in self.column_names[col-1]:
                total_cell.number_format = '#,##0.00'  # Денежные суммы с 2 знаками
            elif '%' in self.column_names[col-1]:
                total_cell.number_format = '0.00'  # Проценты с 2 знаками
            elif any(keyword in self.column_names[col-1].lower() for keyword in ['штук', 'шт', 'кол-во']):
                total_cell.number_format = '#,##0'  # Целые числа для количества
            else:
                total_cell.number_format = '0.00'  # Все остальные числа с 2 знаками
        
        # Ширина колонок
        for col in range(1, 81):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 15
        
        # Первая колонка шире для названий
        ws.column_dimensions['A'].width = 40
        
        # Закрепление областей
        ws.freeze_panes = 'A4'
    
    def apply_data_formatting(self, ws, num_products):
        """Применение форматирования числовых данных товаров"""
        
        # Форматируем строки с данными товаров (строки 4 и далее)
        for row in range(4, 4 + num_products):
            for col in range(1, 81):  # Все 80 столбцов
                cell = ws.cell(row=row, column=col)
                
                # Определяем тип данных по названию столбца
                col_name = self.column_names[col-1] if col <= len(self.column_names) else ''
                
                if col == 1:  # Наименование - текст
                    cell.number_format = 'General'
                elif col == 2:  # Артикул - текст
                    cell.number_format = 'General'
                elif 'расход' in col_name.lower() or '₽' in col_name:
                    cell.number_format = '#,##0.00'  # Денежные суммы с 2 знаками
                elif '%' in col_name:
                    cell.number_format = '0.00'  # Проценты с 2 знаками
                elif any(keyword in col_name.lower() for keyword in ['штук', 'шт', 'кол-во']):
                    cell.number_format = '#,##0'  # Целые числа для количества
                else:
                    cell.number_format = '0.00'  # Все остальные числа с 2 знаками

    def generate_report(self, analytics_file, sku_file) -> Optional[bytes]:
        """
        Главная функция генерации отчета DRR
        
        Args:
            analytics_file: путь к файлу analytics_report_*.xlsx  
            sku_file: путь к файлу sku_statistics_*.xlsx
            
        Returns:
            bytes: Excel файл в виде байт или None при ошибке
        """
        try:
            logger.info("🚀 ГЕНЕРАЦИЯ ОТЧЕТА DRR - НАЧАЛО")
            
            # 1. Загружаем данные
            df_analytics, df_sku = self.load_source_data(analytics_file, sku_file)
            if df_analytics is None or df_sku is None:
                return None
                
            # 2. Создаем маппинг данных
            product_data = self.create_product_data_mapping(df_analytics, df_sku)
            
            # 3. Создаем Excel
            wb = self.create_excel_with_exact_structure(product_data)
            
            # 4. Сохраняем в BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info("✅ ОТЧЕТ DRR УСПЕШНО СГЕНЕРИРОВАН!")
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"❌ КРИТИЧЕСКАЯ ОШИБКА: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None