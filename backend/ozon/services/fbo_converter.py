import os
import requests
import openpyxl
from io import BytesIO
from datetime import datetime

import logging

logger = logging.getLogger(__name__)


BASE_URL = "https://api-seller.ozon.ru"

TURNOVER_GRADE_MAP = {
    "DEFICIT": "Дефицитный",
    "POPULAR": "Популярный",
    "ACTUAL": "Актуальный",
    "SURPLUS": "Избыточный",
    "NO_SALES": "Без продаж",
    "WAS_NO_SALES": "Был без продаж",
    "RESTRICTED_NO_SALES": "Без продаж, ограничен",
    "COLLECTING_DATA": "Сбор данных",
    "WAITING_FOR_SUPPLY": "Ожидаем поставки",
    "WAS_DEFICIT": "Был дефицитным",
    "WAS_POPULAR": "Был популярным",
    "WAS_ACTUAL": "Был актуальным",
    "WAS_SURPLUS": "Был избыточным",
    "TURNOVER_GRADE_NONE": "-",
    "UNSPECIFIED": "-",
}

PLACEMENT_ZONE_MAP = {
    "SORT": "Сортируемый товар",
    "NON_SORT": "Несортируемый товар",
    "OVERSIZE": "Крупногабаритный товар",
    "PRODUCTS": "Продукты питания",
    "JEWELRY": "Ювелирные изделия",
    "CLOSED_ZONE": "Закрытая зона",
    "DANGEROUS_GOODS": "Опасный товар",
    "UNRESOLVED": "Неизвестная зона",
    "UNSPECIFIED": "-",
}

PLACEHOLDER_TEXT = "покажем после выбора кластера размещения"


class FboConverter:
    def __init__(self):
        self.api_key = os.getenv('OZON_API_KEY')
        self.client_id = os.getenv('OZON_CLIENT_ID')
        self.headers = {
            "Client-Id": self.client_id,
            "Api-Key": self.api_key,
            "Content-Type": "application/json",
        }

    def parse_input_excel(self, file) -> list:
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        # Найдём строку с заголовками и индексы нужных колонок
        header_row = None
        col_article = col_sku = col_qty = None

        for row in ws.iter_rows():
            headers = {str(cell.value).strip(): cell.column - 1
                       for cell in row if cell.value is not None}
            if 'Артикул' in headers:
                header_row = row[0].row
                col_article = headers.get('Артикул')
                col_sku = headers.get('SKU')
                col_qty = headers.get('Количество')
                break

        if not header_row:
            raise ValueError(
                "Не найдена колонка «Артикул». "
                "Убедитесь, что в файле есть строка заголовков с колонками: Артикул, SKU, Количество."
            )
        if col_sku is None:
            raise ValueError(
                "Не найдена колонка «SKU». "
                "SKU — это числовой ID товара в Ozon (не артикул). "
                "Добавьте колонку SKU с Ozon ID из Seller Portal."
            )
        if col_qty is None:
            raise ValueError("Не найдена колонка «Количество».")

        items = []
        bad_sku_examples = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            article = row[col_article]
            sku = row[col_sku]
            qty = row[col_qty]
            if article and sku and qty is not None:
                try:
                    items.append({
                        "article": str(article).strip(),
                        "sku": int(sku),
                        "qty": int(qty),
                    })
                except (ValueError, TypeError):
                    if len(bad_sku_examples) < 3:
                        bad_sku_examples.append(str(sku))

        if not items and bad_sku_examples:
            raise ValueError(
                f"Колонка «SKU» содержит текстовые значения ({', '.join(bad_sku_examples)}…), "
                "а должна содержать числовой Ozon ID товара (например 186234567). "
                "Его можно найти в Ozon Seller Portal в карточке товара."
            )

        return items

    def fetch_product_info(self, offer_ids: list) -> dict:
        """Возвращает dict: offer_id -> {name, product_id, barcode, is_super}"""
        url = f"{BASE_URL}/v3/product/info/list"
        payload = {"offer_id": offer_ids}

        try:
            response = requests.post(url, json=payload, headers=self.headers, timeout=30)
            response.raise_for_status()
            data = response.json()
        except Exception as e:
            logger.error(f"Ошибка /v3/product/info/list: {e}")
            return {}

        result = {}
        for item in data.get("items", []):
            oid = item.get("offer_id", "")
            barcodes = item.get("barcodes") or []
            result[oid] = {
                "name": item.get("name", ""),
                "product_id": item.get("id", ""),
                "barcode": barcodes[0] if barcodes else "",
                "is_super": "Да" if item.get("is_super") else "Нет",
            }
        return result

    def fetch_liquidity(self, skus: list) -> dict:
        """Возвращает dict: sku -> turnover_grade_label"""
        url = f"{BASE_URL}/v1/analytics/stocks"
        result = {}

        # Батчи по 100
        for i in range(0, len(skus), 100):
            batch = skus[i:i + 100]
            payload = {"skus": [str(s) for s in batch]}
            try:
                response = requests.post(url, json=payload, headers=self.headers, timeout=30)
                response.raise_for_status()
                data = response.json()
                for item in data.get("items", []):
                    sku = item.get("sku")
                    grade = item.get("turnover_grade", "UNSPECIFIED")
                    result[sku] = TURNOVER_GRADE_MAP.get(grade, grade)
            except Exception as e:
                logger.error(f"Ошибка /v1/analytics/stocks (batch {i}): {e}")

        return result

    def fetch_placement_zones(self, skus: list) -> dict:
        """Возвращает dict: sku -> placement_zone_label"""
        url = f"{BASE_URL}/v1/product/placement-zone/info"
        result = {}

        # Батчи по 150
        for i in range(0, len(skus), 150):
            batch = skus[i:i + 150]
            payload = {"skus": [str(s) for s in batch]}
            try:
                response = requests.post(url, json=payload, headers=self.headers, timeout=30)
                response.raise_for_status()
                data = response.json()
                for item in data.get("products_placement", []):
                    sku = item.get("sku")
                    zone = item.get("placement_zone", "UNSPECIFIED")
                    result[sku] = PLACEMENT_ZONE_MAP.get(zone, zone)
            except Exception as e:
                logger.error(f"Ошибка /v1/product/placement-zone/info (batch {i}): {e}")

        return result

    def generate_output_excel(self, rows: list) -> BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FBO поставка"

        headers = [
            "артикул",
            "имя (необязательно)",
            "количество",
            "ozon id",
            "штрихкод",
            "признак Super",
            "ликвидность",
            "зона размещения",
            'требуется маркировка "честный знак"',
            'требуется "эВСД"',
        ]
        ws.append(headers)

        for row in rows:
            ws.append([
                row.get("article", ""),
                row.get("name", ""),
                row.get("qty", 0),
                row.get("product_id", ""),
                row.get("barcode", ""),
                row.get("is_super", "Нет"),
                row.get("liquidity", "-"),
                row.get("placement_zone", "-"),
                PLACEHOLDER_TEXT,
                PLACEHOLDER_TEXT,
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def convert(self, file) -> BytesIO:
        items = self.parse_input_excel(file)
        if not items:
            raise ValueError("Входной файл не содержит данных")

        offer_ids = [item["article"] for item in items]
        skus = [item["sku"] for item in items]

        product_info = self.fetch_product_info(offer_ids)
        liquidity = self.fetch_liquidity(skus)
        placement_zones = self.fetch_placement_zones(skus)

        rows = []
        for item in items:
            info = product_info.get(item["article"], {})
            sku = item["sku"]
            rows.append({
                "article": item["article"],
                "name": info.get("name", ""),
                "qty": item["qty"],
                "product_id": info.get("product_id", ""),
                "barcode": info.get("barcode", ""),
                "is_super": info.get("is_super", "Нет"),
                "liquidity": liquidity.get(sku, "-"),
                "placement_zone": placement_zones.get(sku, "-"),
            })

        return self.generate_output_excel(rows)
